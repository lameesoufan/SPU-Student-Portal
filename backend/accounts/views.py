import logging
import os
import re
import csv
import io as io_module

import requests
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.decorators import throttle_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken
from openpyxl import load_workbook
from django.contrib.auth import authenticate
from django.conf import settings
from django.core.mail import send_mail
from django.contrib.auth.hashers import make_password, check_password
from django.utils import timezone
from datetime import timedelta
import secrets
from .models import User, DEPARTMENTS, StudentReference, PasswordResetCode
from .permissions import IsDeanOrAdmin
from .selectors import get_doctors
from .throttles import (
    RegisterRateThrottle, 
    PasswordResetThrottle,
    StudentLoginRequestThrottle,
    StudentLoginVerifyThrottle,
)
from .services import (
    create_user_from_import, change_user_password, assign_hod,
    lookup_student_in_reference, register_verified_student,
    change_user_username, generate_username_suggestions,
)


logger = logging.getLogger(__name__)

ALLOWED_IMPORT_EXTENSIONS = ('.xlsx', '.xlsm', '.xltx', '.xltm')
MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000

# ── Upload Reference (Student Reference DB) ─────────────────────────────────
ALLOWED_REFERENCE_EXTENSIONS = ('.xlsx', '.xls', '.csv')
MAX_REFERENCE_FILE_SIZE = 10 * 1024 * 1024
MAX_REFERENCE_ROWS = 10000


def _set_cookie(response, name, value, max_age, secure=False):
    """Helper to set HttpOnly JWT cookies."""
    response.set_cookie(
        name,
        value,
        max_age=max_age,
        httponly=True,
        secure=secure,
        samesite='Lax',
        path='/',
    )


def _clear_cookie(response, name):
    """Helper to clear a cookie."""
    response.delete_cookie(name, path='/')


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([PasswordResetThrottle])
def request_password_reset(request):
    identifier = str(request.data.get('identifier', '')).strip()
    generic = {'message': 'إذا كان اسم المستخدم صحيحًا فسيصل رمز التحقق إلى البريد الإلكتروني المرتبط بالحساب.'}
    if not identifier:
        return Response({'error': 'أدخل اسم المستخدم.'}, status=400)

    # The user only enters their username. The destination email is always
    # read from the account record in the database; it is never supplied by
    # the unauthenticated client.
    user = User.objects.filter(username__iexact=identifier).first()
    if not user or not user.email:
        return Response(generic)

    PasswordResetCode.objects.filter(user=user, is_used=False).update(is_used=True)
    code = f'{secrets.randbelow(1000000):06d}'
    session_token = secrets.token_urlsafe(48)
    PasswordResetCode.objects.create(
        user=user, code_hash=make_password(code), session_token=session_token,
        expires_at=timezone.now() + timedelta(minutes=10),
    )
    try:
        send_mail(
            'رمز إعادة تعيين كلمة المرور - بوابة SPU',
            f'رمز التحقق الخاص بك هو: {code}\nصلاحية الرمز 10 دقائق. لا تشاركه مع أي شخص.',
            settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False,
        )
    except Exception:
        logger.exception('Password reset email failed for user %s', user.pk)
        return Response({'error': 'تعذر إرسال البريد الإلكتروني حاليًا. تحقق من إعدادات البريد وحاول لاحقًا.'}, status=503)

    return Response({**generic, 'session_token': session_token, 'email_hint': _mask_email(user.email), 'expires_in_seconds': 600})


def _mask_email(email):
    local, _, domain = (email or '').partition('@')
    if not domain:
        return ''
    visible = local[:2] if len(local) > 2 else local[:1]
    return f'{visible}{"*" * max(2, len(local)-len(visible))}@{domain}'


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([PasswordResetThrottle])
def verify_password_reset_code(request):
    session_token = str(request.data.get('session_token', '')).strip()
    code = str(request.data.get('code', '')).strip()
    try:
        reset = PasswordResetCode.objects.select_related('user').get(session_token=session_token, is_used=False)
    except PasswordResetCode.DoesNotExist:
        return Response({'error': 'جلسة الاستعادة غير صالحة أو منتهية.'}, status=400)
    if reset.is_expired():
        reset.is_used = True; reset.save(update_fields=['is_used'])
        return Response({'error': 'انتهت صلاحية الرمز. اطلب رمزًا جديدًا.'}, status=400)
    if reset.failed_attempts >= 5:
        reset.is_used = True; reset.save(update_fields=['is_used'])
        return Response({'error': 'تم تجاوز عدد المحاولات المسموح. اطلب رمزًا جديدًا.'}, status=429)
    if not check_password(code, reset.code_hash):
        reset.failed_attempts += 1; reset.save(update_fields=['failed_attempts'])
        return Response({'error': 'رمز التحقق غير صحيح.'}, status=400)
    return Response({'message': 'تم التحقق من الرمز بنجاح.', 'verified': True})


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([PasswordResetThrottle])
def reset_password_with_code(request):
    session_token = str(request.data.get('session_token', '')).strip()
    code = str(request.data.get('code', '')).strip()
    new_password = str(request.data.get('new_password', ''))
    confirm_password = str(request.data.get('confirm_password', ''))
    if new_password != confirm_password:
        return Response({'error': 'كلمتا المرور غير متطابقتين.'}, status=400)
    try:
        reset = PasswordResetCode.objects.select_related('user').get(session_token=session_token, is_used=False)
    except PasswordResetCode.DoesNotExist:
        return Response({'error': 'جلسة الاستعادة غير صالحة أو منتهية.'}, status=400)
    if reset.is_expired() or not check_password(code, reset.code_hash):
        return Response({'error': 'رمز التحقق غير صحيح أو منتهي الصلاحية.'}, status=400)
    result = change_user_password(user=reset.user, new_password=new_password)
    if not result['ok']:
        return Response({'error': result['error']}, status=400)
    reset.is_used = True; reset.save(update_fields=['is_used'])
    PasswordResetCode.objects.filter(user=reset.user, is_used=False).update(is_used=True)
    return Response({'message': 'تم تغيير كلمة المرور بنجاح. يمكنك تسجيل الدخول الآن.'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    current_password = request.data.get('current_password', '')
    new_password     = request.data.get('new_password', '')
    confirm_password = request.data.get('confirm_password', '')
    if not new_password or not confirm_password:
        return Response({'error': 'كلمة المرور الجديدة وتأكيدها مطلوبان.'}, status=400)
    if not request.user.must_change_password:
        if not current_password or not request.user.check_password(current_password):
            return Response({'error': 'كلمة المرور الحالية غير صحيحة.'}, status=400)
    if new_password != confirm_password:
        return Response({'error': 'كلمتا المرور غير متطابقتين.'}, status=400)
    result = change_user_password(user=request.user, new_password=new_password)
    if not result['ok']:
        return Response({'error': result['error']}, status=400)
    return Response({'message': 'Password changed successfully.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def username_suggestions(request):
    """Return username suggestions for the current user (first-login flow)."""
    suggestions = generate_username_suggestions(user=request.user)
    return Response({
        'suggestions': suggestions,
        'current_username': request.user.username,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_username(request):
    """Change the authenticated user's username. Only allowed once."""
    new_username = request.data.get('new_username', '').strip()
    if not new_username:
        return Response({'error': 'New username is required.'}, status=400)

    result = change_user_username(user=request.user, new_username=new_username)
    if not result['ok']:
        return Response({'error': result['error']}, status=400)

    return Response({
        'message': 'Username changed successfully.',
        'new_username': result['new_username'],
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsDeanOrAdmin])
@parser_classes([MultiPartParser, FormParser])
def import_users(request):
    """استيراد المستخدمين من ملف Excel - نسخة محسّنة مع تقارير أخطاء تفصيلية."""
    if 'file' not in request.FILES:
        return Response({'error': 'File is required.'}, status=400)

    role = request.data.get('role') or request.POST.get('role')
    if not role:
        return Response({'error': 'Role is required.'}, status=400)

    upload = request.FILES['file']
    filename = str(upload.name or '').lower()
    if not filename.endswith(ALLOWED_IMPORT_EXTENSIONS):
        return Response({'error': 'Only Excel files are allowed.'}, status=400)
    if upload.size > MAX_IMPORT_FILE_SIZE:
        return Response({'error': 'File is too large. Maximum size is 10 MB.'}, status=400)

    role = role.lower()
    if role not in ['student', 'doctor']:
        return Response({'error': 'Invalid role.'}, status=400)

    try:
        wb = load_workbook(filename=upload, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.exception('Failed to open Excel file')
        return Response({'error': f'Invalid Excel file: {str(exc)}'}, status=400)

    created_users = []
    errors = []

    try:
        if ws.max_row is not None and ws.max_row > MAX_IMPORT_ROWS + 1:
            return Response(
                {'error': f'File has too many rows. Maximum allowed rows is {MAX_IMPORT_ROWS}.'},
                status=400
            )

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue

            full_name  = row[0] if len(row) > 0 else None
            identifier = row[1] if len(row) > 1 else None
            email      = row[2] if len(row) > 2 else None
            department = row[3] if len(row) > 3 else ''

            if identifier is None:
                continue
            if isinstance(identifier, float) and identifier.is_integer():
                identifier = int(identifier)
            username = str(identifier).strip()
            if not username:
                continue

            result = create_user_from_import(
                username=username,
                email=email,
                role=role,
                password=username,
                department=department,
            )

            if not result.get('ok'):
                errors.append({
                    'row': row_idx,
                    'username': username,
                    'error': result.get('error', 'Unknown error'),
                })
                continue

            user = result['user']
            if full_name:
                try:
                    user.first_name = str(full_name)[:150]
                    user.save(update_fields=['first_name'])
                except Exception as e:
                    errors.append({
                        'row': row_idx,
                        'username': username,
                        'error': f'Created but failed to save name: {str(e)}',
                    })

            created_users.append({'username': username})

    except Exception as exc:
        logger.exception('Import users failed')
        return Response({'error': f'Import failed: {str(exc)}'}, status=400)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # لو ما في أي مستخدم تم إنشاؤه - فيه أخطاء فقط
    if not created_users and errors:
        return Response({
            'error': 'No users were imported. See details below.',
            'details': errors,
        }, status=400)

    # لو فيه نجاحات وفيه أخطاء
    if errors:
        return Response({
            'message': f'{len(created_users)} {role}(s) created successfully, but {len(errors)} row(s) had errors.',
            'users': created_users,
            'errors': errors,
        }, status=200)

    # كلشي تمام
    return Response({
        'message': f'{len(created_users)} {role}(s) created successfully.',
        'users': created_users,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsDeanOrAdmin])
def list_doctors(request):
    return Response(get_doctors())


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsDeanOrAdmin])
def list_departments(request):
    from .selectors import get_hod_by_department
    result = []
    for key, label in DEPARTMENTS:
        hod = get_hod_by_department(key)
        result.append({
            'key': key, 'label': label,
            'hod': {
                'id': hod.id, 'username': hod.username,
                'full_name': f"{hod.first_name} {hod.last_name}".strip() or hod.username,
            } if hod else None,
        })
    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsDeanOrAdmin])
def assign_hod_view(request):
    doctor_id  = request.data.get('doctor_id')
    department = request.data.get('department')
    if not doctor_id or not department:
        return Response({'error': 'doctor_id and department are required.'}, status=400)
    result = assign_hod(doctor_id=doctor_id, department=department)
    if not result['ok']:
        return Response({'error': result['error']}, status=400)
    u = result['user']
    return Response({'message': f'{u.first_name or u.username} assigned as HoD of {department}.', 'user': {
        'id': u.id, 'username': u.username,
        'full_name': f"{u.first_name} {u.last_name}".strip() or u.username,
        'department': u.department,
    }})


# ── Student Self-Registration via External API ──────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([RegisterRateThrottle])
def student_self_register(request):
    """
    Student enters university_id + password.
    Backend verifies against external university API.
    On success: creates local account (if new) and returns JWT.
    """
    university_id = str(request.data.get('university_id', '')).strip()
    password      = str(request.data.get('password', '')).strip()

    if not university_id or not password:
        return Response({'error': 'University ID and password are required.'}, status=400)

    # Step 1: verify against external API
    lookup = lookup_student_in_reference(university_id, password)
    if not lookup['ok']:
        return Response({'error': lookup['error']}, status=403)

    # Step 2: create local account if not exists, otherwise reuse
    result = register_verified_student(university_id=university_id, ref_data=lookup['data'])
    if not result['ok']:
        user = authenticate(username=university_id, password=password)
        if not user:
            return Response({'error': result['error']}, status=409)
    else:
        user = result['user']

    # Step 3: issue JWT as HttpOnly cookies
    refresh = RefreshToken.for_user(user)
    refresh['role']                 = user.role
    refresh['username']             = user.username
    refresh['must_change_password'] = user.must_change_password
    refresh['must_change_username'] = user.must_change_username
    refresh['department']           = user.department

    access_token = str(refresh.access_token)
    refresh_token_str = str(refresh)

    response = Response({
        'message': f'Welcome, {user.first_name or user.username}.',
        'username': user.username,
        'role': user.role,
        'must_change_password': user.must_change_password,
        'must_change_username': user.must_change_username,
        'department': user.department,
    })

    secure = getattr(settings, 'JWT_COOKIE_SECURE', not settings.DEBUG)
    _set_cookie(response, 'access_token', access_token,
                getattr(settings, 'JWT_COOKIE_ACCESS_MAX_AGE', 86400), secure=secure)
    _set_cookie(response, 'refresh_token', refresh_token_str,
                getattr(settings, 'JWT_COOKIE_REFRESH_MAX_AGE', 604800), secure=secure)
    return response


# ── Upload Student Reference Database ────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsDeanOrAdmin])
@parser_classes([MultiPartParser, FormParser])
def upload_reference(request):
    """
    رفع قاعدة بيانات الطلاب المرجعية.
    يقبل: .xlsx, .xls, .csv
    الأعمدة المطلوبة: university_id, email
    الأعمدة الاختيارية: full_name, department, password
    
    ملاحظة: إذا كانت password فارغة، سيستخدم النظام university_id كـ password افتراضي
    """
    logger.info("DEBUG UPLOAD: Function called by user %s", request.user.username)
    logger.info("DEBUG UPLOAD: FILES keys: %s", list(request.FILES.keys()))
    
    if 'file' not in request.FILES:
        return Response({'error': 'File is required.'}, status=400)

    upload = request.FILES['file']
    filename = str(upload.name or '').lower()

    if not filename.endswith(ALLOWED_REFERENCE_EXTENSIONS):
        return Response({
            'error': f'Unsupported file type. Allowed: {", ".join(ALLOWED_REFERENCE_EXTENSIONS)}'
        }, status=400)

    if upload.size > MAX_REFERENCE_FILE_SIZE:
        return Response({'error': 'File is too large. Maximum size is 10 MB.'}, status=400)

    # تحديد نوع الملف
    if filename.endswith('.csv'):
        records, parse_errors = _parse_csv_reference(upload)
    else:
        records, parse_errors = _parse_excel_reference(upload)

    logger.info(f"DEBUG UPLOAD: Parsed {len(records)} records, {len(parse_errors)} parse errors")
    logger.info(f"DEBUG UPLOAD: First record: {records[0] if records else 'NONE'}")

    if parse_errors and not records:
        return Response({
            'error': 'Failed to parse file.',
            'details': parse_errors[:20],
        }, status=400)

    if len(records) > MAX_REFERENCE_ROWS:
        return Response({
            'error': f'File has too many rows. Maximum allowed is {MAX_REFERENCE_ROWS}.'
        }, status=400)

    logger.info(f"DEBUG UPLOAD: Starting to save {len(records)} records...")

    # إدخال البيانات بشكل جماعي (update_or_create)
    created_count = 0
    updated_count = 0
    row_errors = []

    for idx, record in enumerate(records, start=2):
        university_id = str(record.get('university_id', '')).strip()
        if not university_id:
            row_errors.append({'row': idx, 'error': 'Missing university_id'})
            continue

        # إذا كلمة المرور فارغة، استخدم الرقم الجامعي كـ password افتراضي
        raw_password = str(record.get('password', '')).strip()
        # SECURITY: hash on write; never store plain text.
        from django.contrib.auth.hashers import make_password
        password = make_password(raw_password) if raw_password else ''
        if not password:
            password = university_id

        logger.info(f"DEBUG: Processing row {idx}, university_id={university_id}, password={password}")

        try:
            obj, created = StudentReference.objects.update_or_create(
                university_id=university_id,
                defaults={
                    'full_name':  str(record.get('full_name', '')).strip(),
                    'department': str(record.get('department', '')).strip(),
                    'email':      str(record.get('email', '')).strip(),
                    'password':   password,
                    'uploaded_by': request.user,
                },
            )
            logger.info(f"DEBUG: Saved {university_id}, created={created}, obj.id={obj.id}")
            if created:
                created_count += 1
            else:
                updated_count += 1
        except Exception as e:
            logger.exception(f"DEBUG: Error saving {university_id}")
            row_errors.append({
                'row': idx,
                'university_id': university_id,
                'error': str(e),
            })

    logger.info(
        'Reference upload by %s: %d created, %d updated, %d errors',
        request.user.username, created_count, updated_count, len(row_errors)
    )

    return Response({
        'message': f'Reference database updated. {created_count} new, {updated_count} updated.',
        'count': created_count + updated_count,
        'created': created_count,
        'updated': updated_count,
        'errors': row_errors[:20],
        'errors_count': len(row_errors),
    }, status=200)


def _parse_excel_reference(upload):
    """
    تحليل ملف Excel.
    الأعمدة المطلوبة: university_id, email
    الأعمدة الاختيارية: full_name, department, password
    إذا password فارغة = university_id
    """
    records = []
    errors = []
    try:
        wb = load_workbook(filename=upload, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.exception('Failed to open Excel reference file')
        return [], [f'Invalid Excel file: {str(exc)}']

    try:
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return [], ['File is empty.']

        # اكتشاف الـ header
        header = rows[0]
        col_map = _detect_columns(header)

        if 'university_id' not in col_map:
            return [], ['Could not find university_id column. Expected headers: university_id, full_name, department, email, password']

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row:
                continue
            record = {}
            for field, col_idx in col_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if isinstance(val, float) and val.is_integer():
                        val = str(int(val))
                    record[field] = str(val) if val is not None else ''
            records.append(record)
    except Exception as exc:
        errors.append(f'Error reading rows: {str(exc)}')
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return records, errors


def _parse_csv_reference(upload):
    """تحليل ملف CSV"""
    records = []
    errors = []
    try:
        # قراءة الملف كنص
        decoded = upload.read().decode('utf-8-sig')  # utf-8-sig يتجاهل BOM
        reader = csv.DictReader(io_module.StringIO(decoded))
        for row_idx, row in enumerate(reader, start=2):
            # تطبيع أسماء الأعمدة (حروف صغيرة، إزالة المسافات)
            normalized = {k.lower().strip(): v for k, v in row.items() if k}
            record = {
                'university_id': normalized.get('university_id', ''),
                'full_name':     normalized.get('full_name', ''),
                'department':    normalized.get('department', ''),
                'email':         normalized.get('email', ''),
                'password':      normalized.get('password', ''),
            }
            records.append(record)
    except Exception as exc:
        errors.append(f'CSV parse error: {str(exc)}')

    return records, errors


def _detect_columns(header):
    """تحويل header row إلى خريطة {field_name: column_index}"""
    col_map = {}
    aliases = {
        'university_id': ['university_id', 'universityid', 'id', 'student_id', 'studentid', 'رقم الجامعي', 'الرقم الجامعي'],
        'full_name':     ['full_name', 'fullname', 'name', 'الاسم', 'اسم الطالب'],
        'department':    ['department', 'dept', 'القسم'],
        'email':         ['email', 'البريد'],
        'password':      ['password', 'pwd', 'pass', 'كلمة المرور', 'الرقم السري'],
    }
    for col_idx, cell in enumerate(header or []):
        if cell is None:
            continue
        cell_lower = str(cell).lower().strip()
        for field, alias_list in aliases.items():
            if cell_lower in [a.lower() for a in alias_list]:
                col_map[field] = col_idx
                break
    return col_map



# ── Student OTP Login (2FA) ─────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
# @throttle_classes([StudentLoginRequestThrottle])  # تعطيل مؤقت للتطوير
def student_login_request(request):
    """
    Step 1 of 2FA: Verify credentials and send OTP.
    
    Request body:
        {
            "university_id": str,
            "password": str
        }
    
    Response (200):
        {
            "message": "OTP sent to your email",
            "session_token": str,
            "email_hint": "xxx...@student.spu.edu",
            "expires_in_seconds": 600
        }
    
    Response (400/403):
        {"error": str}
    """
    from .services import generate_otp, send_otp_email
    from django.contrib.auth import authenticate
    
    university_id = str(request.data.get('university_id', '')).strip()
    password = str(request.data.get('password', '')).strip()
    
    # Validate input
    if not university_id or not password:
        return Response({'error': 'University ID and password are required.'}, status=400)
    
    # Verify credentials against User table (not StudentReference)
    user = authenticate(username=university_id, password=password)
    if not user or user.role != 'student':
        logger.warning('Failed student login request for %s from IP %s', 
                      university_id, request.META.get('REMOTE_ADDR', 'unknown'))
        return Response({'error': 'Invalid credentials'}, status=403)
    
    # Get student email
    email = user.email
    if not email:
        logger.error('No email found for student %s', university_id)
        return Response({'error': 'Student email not configured. Please contact administration.'}, status=500)

    # After the first successful password change, student login becomes direct.
    # OTP is only used while must_change_password is still required.
    if not user.must_change_password:
        refresh = RefreshToken.for_user(user)
        refresh['role'] = user.role
        refresh['username'] = user.username
        refresh['must_change_password'] = user.must_change_password
        refresh['must_change_username'] = user.must_change_username
        refresh['department'] = user.department or ''

        access_token = str(refresh.access_token)
        refresh_token_str = str(refresh)

        response = Response({
            'message': 'Login successful',
            'access': access_token,
            'username': user.username,
            'role': user.role,
            'must_change_password': user.must_change_password,
            'must_change_username': user.must_change_username,
            'department': user.department or '',
        })

        secure = getattr(settings, 'JWT_COOKIE_SECURE', not settings.DEBUG)
        _set_cookie(response, 'access_token', access_token,
                    getattr(settings, 'JWT_COOKIE_ACCESS_MAX_AGE', 86400), secure=secure)
        _set_cookie(response, 'refresh_token', refresh_token_str,
                    getattr(settings, 'JWT_COOKIE_REFRESH_MAX_AGE', 604800), secure=secure)

        logger.info('Student %s logged in directly after password change', university_id)
        return response
    
    # Generate OTP
    ip_address = request.META.get('REMOTE_ADDR', None)
    otp_result = generate_otp(university_id=university_id, ip_address=ip_address)
    
    if not otp_result['ok']:
        logger.error('Failed to generate OTP for %s', university_id)
        return Response({'error': 'An error occurred. Please try again later.'}, status=500)
    
    # Send OTP email
    full_name = f"{user.first_name} {user.last_name}".strip() or university_id
    email_sent = send_otp_email(
        email=email,
        full_name=full_name,
        otp_code=otp_result['otp_code']
    )
    
    if not email_sent:
        logger.error('Failed to send OTP email for %s', university_id)
        return Response({'error': 'Failed to send OTP email. Please try again later.'}, status=500)
    
    # Mask email for privacy (show first 3 chars and domain)
    email_parts = email.split('@')
    if len(email_parts) == 2:
        email_hint = f"{email_parts[0][:3]}...@{email_parts[1]}"
    else:
        email_hint = 'xxx...@student.spu.edu'
    
    logger.info('OTP sent for student %s to email %s', university_id, email)
    
    return Response({
        'message': 'OTP sent to your email',
        'session_token': otp_result['session_token'],
        'email_hint': email_hint,
        'expires_in_seconds': otp_result['expires_in_seconds'],
    })


@api_view(['POST'])
@permission_classes([AllowAny])
# @throttle_classes([StudentLoginVerifyThrottle])  # تعطيل مؤقت للتطوير
def student_login_verify(request):
    """
    Step 2 of 2FA: Verify OTP and complete login.
    
    Request body:
        {
            "session_token": str,
            "code": str  # 6-digit OTP
        }
    
    Response (200):
        {
            "message": "Login successful",
            "access": str,
            "username": str,
            "role": "student",
            "must_change_password": bool,
            "must_change_username": bool,
            "department": str
        }
        + HttpOnly cookies: access_token, refresh_token
    
    Response (400/403):
        {"error": str, "attempts_remaining": int}
    """
    from .services import verify_otp
    
    session_token = str(request.data.get('session_token', '')).strip()
    code = str(request.data.get('code', '')).strip()
    
    # Validate input
    if not session_token or not code:
        return Response({'error': 'Session token and OTP code are required.'}, status=400)
    
    # Verify OTP
    ip_address = request.META.get('REMOTE_ADDR', None)
    verify_result = verify_otp(session_token=session_token, code=code, ip_address=ip_address)
    
    if not verify_result['ok']:
        error_response = {'error': verify_result['error']}
        if 'attempts_remaining' in verify_result:
            error_response['attempts_remaining'] = verify_result['attempts_remaining']
        return Response(error_response, status=403)
    
    # OTP verified successfully - get user from database
    university_id = verify_result['university_id']
    
    # Get user from User table (not StudentReference)
    try:
        user = User.objects.get(username=university_id, role='student')
    except User.DoesNotExist:
        logger.error('Student %s not found in User table after OTP verification', university_id)
        return Response({'error': 'Student data not found. Please contact administration.'}, status=403)
    
    # Issue JWT tokens
    refresh = RefreshToken.for_user(user)
    refresh['role'] = user.role
    refresh['username'] = user.username
    refresh['must_change_password'] = user.must_change_password
    refresh['must_change_username'] = user.must_change_username
    refresh['department'] = user.department or ''
    
    access_token = str(refresh.access_token)
    refresh_token_str = str(refresh)
    
    response = Response({
        'message': 'Login successful',
        'access': access_token,
        'username': user.username,
        'role': user.role,
        'must_change_password': user.must_change_password,
        'must_change_username': user.must_change_username,
        'department': user.department or '',
    })
    
    # Set HttpOnly cookies
    secure = getattr(settings, 'JWT_COOKIE_SECURE', not settings.DEBUG)
    _set_cookie(response, 'access_token', access_token,
                getattr(settings, 'JWT_COOKIE_ACCESS_MAX_AGE', 86400), secure=secure)
    _set_cookie(response, 'refresh_token', refresh_token_str,
                getattr(settings, 'JWT_COOKIE_REFRESH_MAX_AGE', 604800), secure=secure)
    
    logger.info('Student %s logged in successfully via OTP', university_id)
    
    return response
