import csv
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.core.validators import validate_email
from django.db import transaction
from openpyxl import load_workbook

from project_imports.constants import resolve_header_field


User = get_user_model()


class Command(BaseCommand):
    help = (
        'Update imported student email addresses from a CSV or XLSX file. '
        'Required columns: university_id/الرقم الجامعي and email/البريد الإلكتروني.'
    )

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str)
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Validate and show what would change without saving.',
        )

    def handle(self, *args, **options):
        path = Path(options['file_path']).expanduser().resolve()
        if not path.exists() or not path.is_file():
            raise CommandError(f'File not found: {path}')

        rows = self._read_rows(path)
        if not rows:
            raise CommandError('The file contains no student rows.')

        prepared = []
        errors = []
        seen_ids = set()
        seen_emails = set()

        for row_number, row in rows:
            university_id = str(row.get('university_id', '') or '').strip()
            email = str(row.get('email', '') or '').strip().lower()

            if not university_id and not email:
                continue
            if not university_id:
                errors.append(f'Row {row_number}: university ID is missing.')
                continue
            if not email:
                errors.append(f'Row {row_number}: email is missing for student {university_id}.')
                continue
            try:
                validate_email(email)
            except ValidationError:
                errors.append(f'Row {row_number}: invalid email {email!r}.')
                continue
            if university_id in seen_ids:
                errors.append(f'Row {row_number}: duplicate university ID {university_id}.')
                continue
            if email in seen_emails:
                errors.append(f'Row {row_number}: duplicate email {email}.')
                continue

            seen_ids.add(university_id)
            seen_emails.add(email)
            prepared.append((row_number, university_id, email))

        if errors:
            for message in errors[:50]:
                self.stderr.write(self.style.ERROR(message))
            if len(errors) > 50:
                self.stderr.write(self.style.ERROR(f'... and {len(errors) - 50} more errors.'))
            raise CommandError(f'Validation failed with {len(errors)} error(s). No emails were changed.')

        students = {
            user.username: user
            for user in User.objects.filter(username__in=[item[1] for item in prepared])
        }
        email_owners = {
            user.email.strip().lower(): user
            for user in User.objects.exclude(email='')
            if user.email
        }

        updates = []
        for row_number, university_id, email in prepared:
            student = students.get(university_id)
            if not student:
                errors.append(f'Row {row_number}: student {university_id} does not exist.')
                continue
            if student.role != 'student':
                errors.append(f'Row {row_number}: account {university_id} is not a student.')
                continue
            owner = email_owners.get(email)
            if owner and owner.pk != student.pk:
                errors.append(f'Row {row_number}: email {email} is already used by {owner.username}.')
                continue
            if student.email and student.email.strip().lower() != email:
                errors.append(
                    f'Row {row_number}: student {university_id} already has a different email '
                    f'({student.email}). Use Django admin to change it intentionally.'
                )
                continue
            if student.email.strip().lower() == email:
                continue
            student.email = email
            updates.append(student)

        if errors:
            for message in errors[:50]:
                self.stderr.write(self.style.ERROR(message))
            if len(errors) > 50:
                self.stderr.write(self.style.ERROR(f'... and {len(errors) - 50} more errors.'))
            raise CommandError(f'Validation failed with {len(errors)} error(s). No emails were changed.')

        if options['dry_run']:
            self.stdout.write(self.style.WARNING(f'Dry run: {len(updates)} student email(s) would be updated.'))
            for student in updates[:20]:
                self.stdout.write(f'  {student.username} -> {student.email}')
            if len(updates) > 20:
                self.stdout.write(f'  ... and {len(updates) - 20} more')
            return

        with transaction.atomic():
            User.objects.bulk_update(updates, ['email'])

        self.stdout.write(self.style.SUCCESS(f'Updated {len(updates)} student email(s) successfully.'))

    def _read_rows(self, path):
        suffix = path.suffix.lower()
        if suffix == '.csv':
            return self._read_csv(path)
        if suffix == '.xlsx':
            return self._read_xlsx(path)
        raise CommandError('Unsupported file type. Use .csv or .xlsx.')

    def _map_headers(self, headers):
        positions = {}
        for index, header in enumerate(headers):
            field = resolve_header_field(header)
            if field in {'university_id', 'email'} and field not in positions:
                positions[field] = index
        missing = [field for field in ('university_id', 'email') if field not in positions]
        if missing:
            raise CommandError(
                'Missing required column(s): ' + ', '.join(missing) + '. '
                'Use university_id/الرقم الجامعي and email/البريد الإلكتروني.'
            )
        return positions

    def _read_csv(self, path):
        with path.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.reader(handle)
            try:
                headers = next(reader)
            except StopIteration:
                return []
            positions = self._map_headers(headers)
            rows = []
            for row_number, values in enumerate(reader, start=2):
                row = {
                    field: values[index] if index < len(values) else ''
                    for field, index in positions.items()
                }
                rows.append((row_number, row))
            return rows

    def _read_xlsx(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            iterator = worksheet.iter_rows(values_only=True)
            try:
                headers = [str(value or '') for value in next(iterator)]
            except StopIteration:
                return []
            positions = self._map_headers(headers)
            rows = []
            for row_number, values in enumerate(iterator, start=2):
                values = list(values)
                row = {
                    field: values[index] if index < len(values) else ''
                    for field, index in positions.items()
                }
                rows.append((row_number, row))
            return rows
        finally:
            workbook.close()
