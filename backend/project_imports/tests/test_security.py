"""Security regression tests for bulk project imports."""

import uuid
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook
from rest_framework.test import APIClient

from project_imports.models import ImportRow, ImportSession
from project_imports.name_utils import supervisor_identity_key
from project_imports.services import ImportService, UserMapper, _csv_safe_text
from project_imports.validators import FileValidator, ImportValidationError, ParsedWorkbook


pytestmark = [pytest.mark.django_db, pytest.mark.security]


@pytest.fixture(autouse=True)
def clear_import_cache():
    cache.clear()
    yield
    cache.clear()


def authenticated_client(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def import_result(**overrides):
    data = {
        "import_session_id": None,
        "preview_result_id": str(uuid.uuid4()),
        "file_hash": "a" * 64,
        "dry_run": True,
        "status": "preview",
        "partial_import": False,
        "total_rows_processed": 1,
        "valid_rows_count": 1,
        "invalid_rows_count": 0,
        "successful_imports": 0,
        "failed_imports": 0,
        "created_students_count": 0,
        "created_supervisors_count": 0,
        "created_projects_count": 0,
        "users_to_create": {"students": [], "supervisors": []},
        "projects_to_create": 1,
        "validation_errors": [],
        "warnings": [],
        "errors_by_type": {},
        "execution_time_seconds": 0.01,
        "supervisor_credentials_export": None,
        "student_credentials_export": None,
    }
    data.update(overrides)
    return data


def fake_xlsx(name="projects.xlsx", content=b"PK-security-test"):
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def real_xlsx(*, formula=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "Student Name",
        "University ID",
        "Email",
        "Project Title",
        "Department",
        "Supervisor",
        "Project Type",
        "GitHub Repository",
    ])
    title = "=2+2" if formula else "Secure Project"
    sheet.append([
        "Secure Student",
        "20269901",
        "secure@example.com",
        title,
        "software engineering",
        "Secure Doctor",
        "seasonal",
        "https://github.com/example/secure-project",
    ])
    from io import BytesIO

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return SimpleUploadedFile(
        "projects.xlsx",
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def valid_row(**overrides):
    row = {
        "row_number": 2,
        "project_row_number": 2,
        "is_project_leader": True,
        "student_name": "Secure Student",
        "university_id": "20269901",
        "email": "secure@example.com",
        "title": "Secure Project",
        "department": "software_engineering",
        "supervisor_name": "Secure Doctor",
        "project_type": "seasonal",
        "github_repo": "https://github.com/example/secure-project",
    }
    row.update(overrides)
    return row


def create_session(user, **overrides):
    values = {
        "super_admin": user,
        "filename": "projects.xlsx",
        "file_size_bytes": 1024,
        "total_rows": 1,
        "successful_rows": 1,
        "failed_rows": 0,
        "status": ImportSession.STATUS_SUCCESS,
    }
    values.update(overrides)
    return ImportSession.objects.create(**values)


class TestAuthenticationAndRoleBoundaries:
    @pytest.mark.parametrize(
        ("method", "url_name", "with_session"),
        [
            ("post", "import-projects", False),
            ("get", "import-template", False),
            ("get", "import-history", False),
            ("get", "import-rows", True),
        ],
    )
    def test_anonymous_user_cannot_reach_import_resources(self, api_client, method, url_name, with_session):
        url = reverse(url_name, args=[uuid.uuid4()]) if with_session else reverse(url_name)
        if method == "post":
            response = api_client.post(url, {"file": fake_xlsx()}, format="multipart")
        else:
            response = api_client.get(url)
        assert response.status_code in {401, 403}

    @pytest.mark.parametrize("client_fixture", ["student_client", "doctor_client", "hod_client"])
    def test_non_dean_roles_cannot_start_import(self, request, client_fixture):
        client = request.getfixturevalue(client_fixture)
        response = client.post(reverse("import-projects"), {"file": fake_xlsx()}, format="multipart")
        assert response.status_code == 403


class TestSensitiveResponseCaching:
    @pytest.mark.parametrize("mode", ["preview", "success", "validation_error", "server_error"])
    def test_import_responses_are_never_cacheable(self, dean_client, mode):
        url = reverse("import-projects")
        if mode == "preview":
            side_effect = None
            result = import_result()
            url += "?dry_run=true"
        elif mode == "success":
            side_effect = None
            result = import_result(dry_run=False, status="success", file_hash=None, preview_result_id=None)
        elif mode == "validation_error":
            side_effect = ImportValidationError("Invalid workbook")
            result = None
        else:
            side_effect = RuntimeError("database password=secret")
            result = None

        with patch("project_imports.views.ImportService.execute_import", return_value=result, side_effect=side_effect):
            response = dean_client.post(url, {"file": fake_xlsx()}, format="multipart")

        assert "no-store" in response["Cache-Control"]
        assert "private" in response["Cache-Control"]
        assert response["Pragma"] == "no-cache"
        assert "password=secret" not in str(getattr(response, "data", ""))


class TestPreviewIsolationSecurity:
    def test_preview_cache_contains_only_minimum_binding_metadata(self, dean):
        service = ImportService(dean)
        preview_id = service._cache_preview("a" * 64, 3)
        cached = cache.get(service._preview_key(preview_id))
        assert set(cached) == {"user_id", "file_hash", "valid_rows_count"}
        assert cached == {"user_id": dean.id, "file_hash": "a" * 64, "valid_rows_count": 3}

    def test_preview_created_by_another_dean_cannot_be_reused(self, dean, user_factory):
        other = user_factory(role="dean")
        owner_service = ImportService(dean)
        preview_id = owner_service._cache_preview("a" * 64, 1)
        with pytest.raises(ImportValidationError, match="expired"):
            ImportService(other)._validate_preview("a" * 64, preview_id)

    def test_preview_is_bound_to_exact_file_hash(self, dean):
        service = ImportService(dean)
        preview_id = service._cache_preview("a" * 64, 1)
        with pytest.raises(ImportValidationError, match="does not match"):
            service._validate_preview("b" * 64, preview_id)

    def test_malformed_preview_reference_is_rejected_before_cache_lookup(self, dean):
        service = ImportService(dean)
        with patch("project_imports.services.cache.get") as cache_get:
            with pytest.raises(ImportValidationError, match="Invalid preview reference"):
                service._validate_preview("a" * 64, "../not-a-uuid\n")
        cache_get.assert_not_called()

    def test_expired_preview_reference_is_rejected(self, dean):
        with pytest.raises(ImportValidationError, match="expired"):
            ImportService(dean)._validate_preview("a" * 64, str(uuid.uuid4()))

    def test_valid_preview_reference_is_accepted(self, dean):
        service = ImportService(dean)
        preview_id = service._cache_preview("a" * 64, 1)
        service._validate_preview("a" * 64, preview_id)

    def test_preview_identifiers_are_unique_unpredictable_uuids(self, dean):
        service = ImportService(dean)
        first = service._cache_preview("a" * 64, 1)
        second = service._cache_preview("a" * 64, 1)
        assert uuid.UUID(first)
        assert uuid.UUID(second)
        assert first != second
        assert first != "a" * 64


class TestTemporaryCredentialSecurity:
    def test_same_identifier_never_generates_same_default_password(self, dean):
        mapper = UserMapper()
        first = mapper.generate_password("20269901")
        second = mapper.generate_password("20269901")
        assert first != second
        assert len(first) >= 16
        assert len(second) >= 16

    def test_legacy_deterministic_pattern_gets_random_suffix(self, settings):
        settings.IMPORT_TEMP_PASSWORD_FORMAT = "TEMP-{identifier}-2026!"
        first = UserMapper().generate_password("20269901")
        second = UserMapper().generate_password("20269901")
        assert first.startswith("TEMP-20269901-2026!-")
        assert second.startswith("TEMP-20269901-2026!-")
        assert first != second

    def test_explicit_random_placeholder_is_supported(self, settings):
        settings.IMPORT_TEMP_PASSWORD_FORMAT = "SPU-{identifier}-{random}!"
        password = UserMapper().generate_password("20269901")
        assert password.startswith("SPU-20269901-")
        assert "{random}" not in password

    def test_created_student_password_is_hashed_at_rest(self, settings):
        settings.IMPORT_TEMP_PASSWORD_FORMAT = "TEST-{identifier}-{random}!"
        result = UserMapper().resolve_users([valid_row()])
        student = result["students"]["20269901"]
        raw = result["credentials"][student.id]["password"]
        assert student.password != raw
        assert student.check_password(raw)

    def test_created_supervisor_password_is_hashed_at_rest(self, settings):
        settings.IMPORT_TEMP_PASSWORD_FORMAT = "TEST-{identifier}-{random}!"
        result = UserMapper().resolve_users([valid_row()])
        supervisor = result["supervisors"][2]
        raw = result["credentials"][supervisor.id]["password"]
        assert supervisor.password != raw
        assert supervisor.check_password(raw)

    def test_preview_result_never_contains_plaintext_credential_exports(self, dean):
        service = ImportService(dean)
        parsed = ParsedWorkbook("projects.xlsx", 100, "a" * 64, [valid_row()])
        result = service._build_result(
            parsed=parsed,
            session=None,
            issues=[],
            dry_run=True,
            execution_time=0.01,
            plan={"students_to_create": ["20269901"], "supervisors_to_create": [{"username": "doctor"}]},
            created=None,
            preview_result_id=str(uuid.uuid4()),
        )
        assert result["student_credentials_export"] is None
        assert result["supervisor_credentials_export"] is None
        assert "password" not in str(result["users_to_create"]).lower()


class TestWorkbookInputSecurity:
    def test_legacy_xls_is_rejected(self):
        with pytest.raises(ImportValidationError, match="Legacy .xls"):
            FileValidator().validate_file(SimpleUploadedFile("projects.xls", b"legacy"))

    def test_non_excel_extension_is_rejected(self):
        with pytest.raises(ImportValidationError, match="Expected .xlsx"):
            FileValidator().validate_file(SimpleUploadedFile("projects.csv", b"a,b"))

    def test_oversized_file_is_rejected_before_content_is_read(self):
        upload = SimpleNamespace(name="projects.xlsx", size=(10 * 1024 * 1024) + 1)
        upload.read = lambda: (_ for _ in ()).throw(AssertionError("must not read oversized upload"))
        with pytest.raises(ImportValidationError) as exc:
            FileValidator().validate_file(upload)
        assert exc.value.status_code == 413

    def test_macro_payload_is_rejected(self):
        from io import BytesIO
        import zipfile

        stream = BytesIO()
        with zipfile.ZipFile(stream, "w") as archive:
            archive.writestr("xl/vbaProject.bin", b"macro")
        with pytest.raises(ImportValidationError, match="macros"):
            FileValidator().validate_file(SimpleUploadedFile("projects.xlsx", stream.getvalue()))

    def test_corrupted_xlsx_returns_generic_validation_error(self):
        with pytest.raises(ImportValidationError) as exc:
            FileValidator().parse_workbook(fake_xlsx(content=b"not-an-xlsx"))
        text = exc.value.message.lower()
        assert "corrupted" in text or "valid excel" in text
        assert "traceback" not in text

    def test_formula_in_imported_field_is_rejected(self):
        with pytest.raises(ImportValidationError, match="Formula cells are not allowed"):
            FileValidator().parse_workbook(real_xlsx(formula=True))

    def test_malformed_workbook_creates_no_audit_session(self, dean):
        before = ImportSession.objects.count()
        with pytest.raises(ImportValidationError):
            ImportService(dean).execute_import(fake_xlsx(content=b"broken"), dry_run=False)
        assert ImportSession.objects.count() == before


class TestExecutionFailureAndAuditIsolation:
    def test_internal_execution_exception_is_not_persisted_in_history(self, dean):
        service = ImportService(dean)
        parsed = ParsedWorkbook("projects.xlsx", 100, "a" * 64, [valid_row()])
        with (
            patch.object(service.file_validator, "parse_workbook", return_value=parsed),
            patch.object(service.row_validator, "validate_rows", return_value=([valid_row()], [])),
            patch.object(service.user_mapper, "build_plan", return_value={"issues": [], "students_to_create": [], "supervisors_to_create": []}),
            patch.object(service, "_execute_batched_import", side_effect=RuntimeError("postgres password=TOPSECRET host=/internal")),
        ):
            with pytest.raises(RuntimeError):
                service.execute_import(fake_xlsx(), dry_run=False)

        session = ImportSession.objects.get(super_admin=dean)
        assert session.error_summary == "Import failed during execution."
        assert "TOPSECRET" not in session.error_summary
        assert "/internal" not in session.error_summary

    def test_history_api_does_not_reveal_previous_internal_exception_details(self, dean, dean_client):
        create_session(dean, status=ImportSession.STATUS_FAILED, error_summary="Import failed during execution.")
        response = dean_client.get(reverse("import-history"))
        assert response.status_code == 200
        assert "password=" not in str(response.data).lower()
        assert "traceback" not in str(response.data).lower()

    def test_history_is_isolated_between_super_admins(self, dean, dean_client, user_factory):
        other = user_factory(role="dean")
        own = create_session(dean, filename="own.xlsx")
        foreign = create_session(other, filename="foreign.xlsx")
        response = dean_client.get(reverse("import-history"))
        ids = {row["id"] for row in response.data}
        assert str(own.id) in {str(value) for value in ids}
        assert str(foreign.id) not in {str(value) for value in ids}

    def test_rows_are_isolated_between_super_admins(self, dean_client, user_factory):
        other = user_factory(role="dean")
        session = create_session(other)
        ImportRow.objects.create(session=session, row_number=2, university_id="20269999", project_title="Secret", status="success")
        response = dean_client.get(reverse("import-rows", args=[session.id]))
        assert response.status_code == 200
        assert response.data == []

    def test_history_payload_excludes_admin_account_secrets(self, dean, dean_client):
        create_session(dean)
        response = dean_client.get(reverse("import-history"))
        payload = str(response.data).lower()
        assert "password" not in payload
        assert "email" not in payload
        assert "is_superuser" not in payload

    def test_row_payload_excludes_created_student_account_secrets(self, dean, dean_client, user_factory):
        student = user_factory(role="student", username="20269977", email="private@example.com")
        session = create_session(dean)
        ImportRow.objects.create(
            session=session,
            row_number=2,
            university_id=student.username,
            project_title="Public title",
            status="success",
            created_student=student,
        )
        response = dean_client.get(reverse("import-rows", args=[session.id]))
        payload = str(response.data).lower()
        assert "private@example.com" not in payload
        assert "password" not in payload


class TestCsvFormulaInjectionProtection:
    @pytest.mark.parametrize("dangerous", ["=2+2", "+SUM(A1:A2)", "-10+20", "@cmd"])
    def test_formula_like_export_cells_are_neutralized(self, dangerous):
        safe = _csv_safe_text(dangerous)
        assert safe.startswith("'")
        assert safe[1:] == dangerous

    def test_normal_export_text_is_unchanged(self):
        assert _csv_safe_text("Normal Project") == "Normal Project"

    def test_student_credentials_export_neutralizes_user_controlled_text(self, dean, user_factory):
        student = user_factory(role="student", username="20269901", first_name="=HYPERLINK", last_name="Test")
        service = ImportService(dean)
        row = valid_row(title="=CMD|'/C calc'!A0")
        user_map = {
            "students": {student.username: student},
            "created_students": [student],
            "credentials": {student.id: {"password": "SafePassword!123"}},
        }
        export = service._build_student_credentials_export([row], user_map)
        exported = export["rows"][0]
        assert exported["full_name"].startswith("'")
        assert exported["project_title"].startswith("'")

    def test_supervisor_credentials_export_neutralizes_user_controlled_text(self, dean, user_factory):
        supervisor = user_factory(role="doctor", first_name="@Doctor", last_name="Name")
        service = ImportService(dean)
        row = valid_row(title="+SUM(A1:A2)", supervisor_name=supervisor.get_full_name())
        identity = supervisor_identity_key(supervisor.get_full_name())
        user_map = {
            "credentials": {supervisor.id: {"password": "SafePassword!123"}},
            "created_supervisors": [supervisor],
            "identity_to_user": {identity: supervisor},
            "supervisors": {2: supervisor},
            "co_supervisors_map": {},
        }
        export = service._build_supervisor_credentials_export([row], user_map)
        exported = export["rows"][0]
        assert exported["full_name"].startswith("'")
        assert exported["project_titles"].startswith("'")


class TestImportLockIsolation:
    def test_lock_for_one_dean_does_not_block_another_dean(self, dean, user_factory):
        other = user_factory(role="dean")
        cache.set(f"project_import_in_progress_{dean.id}", True, timeout=3600)
        client = authenticated_client(other)
        with patch("project_imports.views.ImportService.execute_import", return_value=import_result()):
            response = client.post(
                reverse("import-projects") + "?dry_run=true",
                {"file": fake_xlsx()},
                format="multipart",
            )
        assert response.status_code == 200

    def test_same_dean_cannot_bypass_existing_lock(self, dean, dean_client):
        cache.set(f"project_import_in_progress_{dean.id}", True, timeout=3600)
        with patch("project_imports.views.ImportService.execute_import") as execute:
            response = dean_client.post(reverse("import-projects"), {"file": fake_xlsx()}, format="multipart")
        assert response.status_code == 409
        execute.assert_not_called()
