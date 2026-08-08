"""HTTP API tests for project bulk-import preview, execution, template, and audit history."""

from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

import pytest
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from project_imports.models import ImportRow, ImportSession
from project_imports.validators import ImportValidationError


pytestmark = [pytest.mark.django_db, pytest.mark.api]


@pytest.fixture(autouse=True)
def clear_import_cache():
    cache.clear()
    yield
    cache.clear()


def xlsx_upload(name="projects.xlsx", content=b"PK-test-workbook"):
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def import_result(**overrides):
    values = {
        "import_session_id": None,
        "preview_result_id": "preview-123",
        "file_hash": "abc123",
        "dry_run": True,
        "status": "preview",
        "partial_import": False,
        "total_rows_processed": 2,
        "valid_rows_count": 2,
        "invalid_rows_count": 0,
        "successful_imports": 0,
        "failed_imports": 0,
        "created_students_count": 0,
        "created_supervisors_count": 0,
        "created_projects_count": 0,
        "users_to_create": {"students": [], "supervisors": []},
        "projects_to_create": 1,
        "validation_errors": [],
        "warnings": [],
        "errors_by_type": {},
        "execution_time_seconds": 0.01,
        "supervisor_credentials_export": None,
        "student_credentials_export": None,
    }
    values.update(overrides)
    return values


def create_session(user, **overrides):
    values = {
        "super_admin": user,
        "filename": "projects.xlsx",
        "file_size_bytes": 1024,
        "total_rows": 2,
        "successful_rows": 2,
        "failed_rows": 0,
        "status": ImportSession.STATUS_SUCCESS,
        "completed_at": timezone.now(),
    }
    values.update(overrides)
    return ImportSession.objects.create(**values)


def create_row(session, number=2, **overrides):
    values = {
        "session": session,
        "row_number": number,
        "university_id": f"2026{number:04d}",
        "project_title": f"Imported Project {number}",
        "status": ImportRow.STATUS_SUCCESS,
        "error_message": "",
    }
    values.update(overrides)
    return ImportRow.objects.create(**values)


class TestImportProjectsApi:
    def test_missing_file_is_rejected_before_service_call(self, dean_client):
        with patch("project_imports.views.ImportService.execute_import") as execute:
            response = dean_client.post(reverse("import-projects"), {}, format="multipart")

        assert response.status_code == 400
        assert response.data == {"error": "File is required."}
        execute.assert_not_called()

    @pytest.mark.parametrize("value", ["true", "TRUE", "True"])
    def test_query_string_dry_run_is_forwarded_as_boolean(self, value, dean_client):
        result = import_result()
        with patch("project_imports.views.ImportService.execute_import", return_value=result) as execute:
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run={value}',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 200
        assert response.data["dry_run"] is True
        assert execute.call_args.kwargs["dry_run"] is True

    def test_form_dry_run_is_supported(self, dean_client):
        result = import_result()
        with patch("project_imports.views.ImportService.execute_import", return_value=result) as execute:
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload(), "dry_run": "true"},
                format="multipart",
            )

        assert response.status_code == 200
        assert execute.call_args.kwargs["dry_run"] is True

    def test_normal_import_defaults_dry_run_to_false(self, dean_client):
        result = import_result(
            preview_result_id=None,
            file_hash=None,
            dry_run=False,
            status="success",
            import_session_id="session-1",
            successful_imports=1,
            created_projects_count=1,
            projects_to_create=0,
        )
        with patch("project_imports.views.ImportService.execute_import", return_value=result) as execute:
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 201
        assert execute.call_args.kwargs["dry_run"] is False

    def test_preview_result_id_from_form_is_forwarded(self, dean_client):
        result = import_result(dry_run=False, status="success", preview_result_id=None, file_hash=None)
        with patch("project_imports.views.ImportService.execute_import", return_value=result) as execute:
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload(), "preview_result_id": "preview-form"},
                format="multipart",
            )

        assert response.status_code == 201
        assert execute.call_args.kwargs["preview_result_id"] == "preview-form"

    def test_preview_result_id_from_query_is_forwarded(self, dean_client):
        result = import_result(dry_run=False, status="success", preview_result_id=None, file_hash=None)
        with patch("project_imports.views.ImportService.execute_import", return_value=result) as execute:
            response = dean_client.post(
                f'{reverse("import-projects")}?preview_result_id=preview-query',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 201
        assert execute.call_args.kwargs["preview_result_id"] == "preview-query"

    def test_service_is_initialized_with_authenticated_dean(self, dean, dean_client):
        result = import_result()
        with patch("project_imports.views.ImportService") as service_cls:
            service_cls.return_value.execute_import.return_value = result
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run=true',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 200
        service_cls.assert_called_once_with(dean)

    def test_concurrent_import_for_same_user_returns_conflict(self, dean, dean_client):
        lock_key = f"project_import_in_progress_{dean.id}"
        cache.set(lock_key, True, timeout=3600)

        with patch("project_imports.views.ImportService.execute_import") as execute:
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 409
        assert "already in progress" in response.data["error"].lower()
        execute.assert_not_called()

    def test_import_lock_is_released_after_success(self, dean, dean_client):
        lock_key = f"project_import_in_progress_{dean.id}"
        with patch("project_imports.views.ImportService.execute_import", return_value=import_result()):
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run=true',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 200
        assert cache.get(lock_key) is None

    def test_import_lock_is_released_after_validation_error(self, dean, dean_client):
        lock_key = f"project_import_in_progress_{dean.id}"
        error = ImportValidationError("Invalid workbook", details={"field": "file"})
        with patch("project_imports.views.ImportService.execute_import", side_effect=error):
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == error.status_code
        assert response.data == {"error": "Invalid workbook", "details": {"field": "file"}}
        assert cache.get(lock_key) is None

    def test_unexpected_exception_returns_generic_error_and_releases_lock(self, dean, dean_client):
        lock_key = f"project_import_in_progress_{dean.id}"
        with patch(
            "project_imports.views.ImportService.execute_import",
            side_effect=RuntimeError("database password=secret"),
        ):
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 500
        assert response.data == {"error": "Import failed. No changes were saved."}
        assert "secret" not in str(response.data)
        assert cache.get(lock_key) is None

    def test_all_invalid_rows_return_bad_request(self, dean_client):
        result = import_result(
            status="failed",
            valid_rows_count=0,
            invalid_rows_count=2,
            projects_to_create=0,
            validation_errors=[{"row_number": 2}, {"row_number": 3}],
        )
        with patch("project_imports.views.ImportService.execute_import", return_value=result):
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run=true',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 400
        assert response.data["status"] == "failed"

    def test_partial_preview_with_valid_rows_returns_ok(self, dean_client):
        result = import_result(
            status="partial_preview",
            partial_import=True,
            valid_rows_count=1,
            invalid_rows_count=1,
            validation_errors=[{"row_number": 3}],
        )
        with patch("project_imports.views.ImportService.execute_import", return_value=result):
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run=true',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 200
        assert response.data["partial_import"] is True

    def test_partial_real_import_with_valid_rows_returns_created(self, dean_client):
        result = import_result(
            import_session_id="session-1",
            preview_result_id=None,
            file_hash=None,
            dry_run=False,
            status="partial_success",
            partial_import=True,
            valid_rows_count=1,
            invalid_rows_count=1,
            successful_imports=1,
            failed_imports=1,
            created_projects_count=1,
            projects_to_create=0,
            validation_errors=[{"row_number": 3}],
        )
        with patch("project_imports.views.ImportService.execute_import", return_value=result):
            response = dean_client.post(
                reverse("import-projects"),
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 201
        assert response.data["status"] == "partial_success"

    def test_preview_response_preserves_public_result_contract(self, dean_client):
        result = import_result(
            users_to_create={"students": ["20260001"], "supervisors": [{"username": "doctor_x"}]},
            warnings=[{"row_number": 2, "error_type": "duplicate"}],
        )
        with patch("project_imports.views.ImportService.execute_import", return_value=result):
            response = dean_client.post(
                f'{reverse("import-projects")}?dry_run=true',
                {"file": xlsx_upload()},
                format="multipart",
            )

        assert response.status_code == 200
        assert response.data == result
        assert response.data["preview_result_id"] == "preview-123"
        assert response.data["file_hash"] == "abc123"


class TestTemplateDownloadApi:
    def test_dean_downloads_xlsx_template(self, dean_client):
        response = dean_client.get(reverse("import-template"))

        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response["Content-Disposition"] == 'attachment; filename="project_import_template.xlsx"'
        assert response.content.startswith(b"PK")

    def test_generated_template_contains_projects_and_instructions_sheets(self, dean_client):
        response = dean_client.get(reverse("import-template"))
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=False)
        try:
            assert workbook.sheetnames == ["Projects", "Instructions"]
            assert workbook["Projects"].max_row >= 3
            assert workbook["Instructions"].max_row >= 2
        finally:
            workbook.close()

    def test_non_dean_cannot_download_template(self, student_client):
        response = student_client.get(reverse("import-template"))
        assert response.status_code == 403


class TestImportHistoryApi:
    def test_history_returns_only_current_deans_sessions(self, dean, user_factory, dean_client):
        other_dean = user_factory(role="dean", department=None, username="other_dean")
        own = create_session(dean, filename="own.xlsx")
        create_session(other_dean, filename="other.xlsx")

        response = dean_client.get(reverse("import-history"))

        assert response.status_code == 200
        assert [row["id"] for row in response.data] == [str(own.id)]
        assert response.data[0]["filename"] == "own.xlsx"

    def test_history_is_ordered_newest_first(self, dean, dean_client):
        older = create_session(dean, filename="older.xlsx")
        newer = create_session(dean, filename="newer.xlsx")
        ImportSession.objects.filter(pk=older.pk).update(started_at=timezone.now() - timedelta(days=1))

        response = dean_client.get(reverse("import-history"))

        assert response.status_code == 200
        assert [row["id"] for row in response.data] == [str(newer.id), str(older.id)]

    @pytest.mark.parametrize(
        ("status_value", "expected_filename"),
        [
            (ImportSession.STATUS_SUCCESS, "success.xlsx"),
            (ImportSession.STATUS_FAILED, "failed.xlsx"),
            (ImportSession.STATUS_PENDING, "pending.xlsx"),
        ],
    )
    def test_history_filters_by_status(self, status_value, expected_filename, dean, dean_client):
        create_session(dean, filename="success.xlsx", status=ImportSession.STATUS_SUCCESS)
        create_session(dean, filename="failed.xlsx", status=ImportSession.STATUS_FAILED)
        create_session(dean, filename="pending.xlsx", status=ImportSession.STATUS_PENDING, completed_at=None)

        response = dean_client.get(reverse("import-history"), {"status": status_value})

        assert response.status_code == 200
        assert [row["filename"] for row in response.data] == [expected_filename]

    def test_history_filters_by_from_and_to_date(self, dean, dean_client):
        old = create_session(dean, filename="old.xlsx")
        middle = create_session(dean, filename="middle.xlsx")
        recent = create_session(dean, filename="recent.xlsx")
        today = timezone.localdate()
        ImportSession.objects.filter(pk=old.pk).update(started_at=timezone.now() - timedelta(days=10))
        ImportSession.objects.filter(pk=middle.pk).update(started_at=timezone.now() - timedelta(days=5))
        ImportSession.objects.filter(pk=recent.pk).update(started_at=timezone.now() - timedelta(days=1))

        response = dean_client.get(
            reverse("import-history"),
            {
                "from_date": (today - timedelta(days=6)).isoformat(),
                "to_date": (today - timedelta(days=2)).isoformat(),
            },
        )

        assert response.status_code == 200
        assert [row["filename"] for row in response.data] == ["middle.xlsx"]

    def test_invalid_date_filters_are_ignored(self, dean, dean_client):
        session = create_session(dean)

        response = dean_client.get(
            reverse("import-history"),
            {"from_date": "not-a-date", "to_date": "also-bad"},
        )

        assert response.status_code == 200
        assert [row["id"] for row in response.data] == [str(session.id)]

    def test_history_is_not_paginated(self, dean, dean_client):
        for index in range(3):
            create_session(dean, filename=f"history-{index}.xlsx")

        response = dean_client.get(reverse("import-history"))

        assert response.status_code == 200
        assert isinstance(response.data, list)
        assert len(response.data) == 3

    def test_history_payload_contains_declared_read_only_fields(self, dean, dean_client):
        session = create_session(dean, filename="audit.xlsx", error_summary="none")

        response = dean_client.get(reverse("import-history"))

        assert response.status_code == 200
        payload = response.data[0]
        assert set(payload) == {
            "id",
            "super_admin",
            "super_admin_username",
            "filename",
            "file_size_bytes",
            "total_rows",
            "successful_rows",
            "failed_rows",
            "started_at",
            "completed_at",
            "status",
            "error_summary",
        }
        assert payload["id"] == str(session.id)
        assert payload["super_admin"] == dean.id
        assert payload["super_admin_username"] == dean.username


class TestImportRowsApi:
    def test_rows_returns_owned_session_rows_in_model_order(self, dean, dean_client):
        session = create_session(dean)
        second = create_row(session, number=3)
        first = create_row(session, number=2)

        response = dean_client.get(reverse("import-rows", args=[session.id]))

        assert response.status_code == 200
        assert [row["id"] for row in response.data] == [first.id, second.id]
        assert [row["row_number"] for row in response.data] == [2, 3]

    def test_rows_for_other_deans_session_returns_empty_list(self, user_factory, dean_client):
        other_dean = user_factory(role="dean", department=None, username="rows_other_dean")
        session = create_session(other_dean)
        create_row(session)

        response = dean_client.get(reverse("import-rows", args=[session.id]))

        assert response.status_code == 200
        assert response.data == []

    def test_unknown_session_returns_empty_list(self, dean_client):
        import uuid

        response = dean_client.get(reverse("import-rows", args=[uuid.uuid4()]))

        assert response.status_code == 200
        assert response.data == []

    def test_rows_are_not_paginated(self, dean, dean_client):
        session = create_session(dean)
        for number in range(2, 7):
            create_row(session, number=number)

        response = dean_client.get(reverse("import-rows", args=[session.id]))

        assert response.status_code == 200
        assert isinstance(response.data, list)
        assert len(response.data) == 5

    def test_row_payload_exposes_audit_fields_and_public_related_labels(
        self,
        dean,
        user_factory,
        dean_client,
    ):
        from projects.models import StudentIdeaProposal

        student = user_factory(role="student", department="software_engineering", username="imported_student")
        supervisor = user_factory(role="doctor", department="software_engineering", username="import_supervisor")
        proposal = StudentIdeaProposal.objects.create(
            student=student,
            supervisor=supervisor,
            title="Imported API Project",
            description="Imported project for row API.",
            department="software_engineering",
            project_type="seasonal",
            status="assigned",
        )
        session = create_session(dean)
        row = create_row(
            session,
            created_student=student,
            created_project=proposal,
            project_title=proposal.title,
        )

        response = dean_client.get(reverse("import-rows", args=[session.id]))

        assert response.status_code == 200
        payload = response.data[0]
        assert set(payload) == {
            "id",
            "session",
            "row_number",
            "university_id",
            "project_title",
            "status",
            "error_message",
            "created_student",
            "created_student_username",
            "created_project",
            "created_project_title",
        }
        assert payload["id"] == row.id
        assert str(payload["session"]) == str(session.id)
        assert payload["created_student"] == student.id
        assert payload["created_student_username"] == student.username
        assert payload["created_project"] == proposal.id
        assert payload["created_project_title"] == proposal.title

    def test_row_with_deleted_relations_omits_nested_display_fields(self, dean, user_factory, dean_client):
        from projects.models import StudentIdeaProposal

        student = user_factory(role="student", department="software_engineering", username="deleted_import_student")
        supervisor = user_factory(role="doctor", department="software_engineering", username="deleted_import_supervisor")
        proposal = StudentIdeaProposal.objects.create(
            student=student,
            supervisor=supervisor,
            title="Deleted Related Project",
            description="Temporary imported project.",
            department="software_engineering",
            project_type="seasonal",
            status="assigned",
        )
        session = create_session(dean)
        row = create_row(session, created_student=student, created_project=proposal)
        proposal.delete()
        student.delete()
        row.refresh_from_db()

        response = dean_client.get(reverse("import-rows", args=[session.id]))

        assert response.status_code == 200
        payload = response.data[0]
        assert payload["created_student"] is None
        assert payload["created_project"] is None
        assert "created_student_username" not in payload
        assert "created_project_title" not in payload

    def test_non_dean_cannot_read_import_rows(self, student_client):
        import uuid

        response = student_client.get(reverse("import-rows", args=[uuid.uuid4()]))
        assert response.status_code == 403
