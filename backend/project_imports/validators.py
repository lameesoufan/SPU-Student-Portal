import hashlib
import html
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from urllib.parse import urlparse

from django.contrib.auth import get_user_model
from django.db.models import Q
from openpyxl import load_workbook

from projects.models import (
    IdeaApplication,
    ProjectApplication,
    ProposalInvitation,
    StudentIdeaProposal,
    TeamInvitation,
)

from .constants import (
    HEADER_TO_FIELD,
    MAX_FILE_SIZE_BYTES,
    MAX_ROWS,
    REQUIRED_HEADERS,
    VALID_DEPARTMENTS,
    VALID_PROJECT_TYPES,
)


User = get_user_model()


class ImportValidationError(Exception):
    def __init__(self, message, *, status_code=400, details=None):
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.details = details or []


@dataclass
class ValidationIssue:
    row_number: int | None
    field_name: str
    error_message: str
    row_data: dict = field(default_factory=dict)
    level: str = 'error'
    error_type: str = 'validation'

    def to_dict(self):
        return {
            'row_number': self.row_number,
            'field_name': self.field_name,
            'error_message': html.escape(str(self.error_message))[:200],
            'row_data': self.row_data,
            'level': self.level,
            'error_type': self.error_type,
        }


@dataclass
class ParsedWorkbook:
    filename: str
    file_size_bytes: int
    file_hash: str
    rows: list[dict]


def normalize_cell_value(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def sanitize_text(value):
    return normalize_cell_value(value).strip()


class FileValidator:
    allowed_extensions = ('.xlsx',)

    def validate_file(self, upload) -> bytes:
        filename = str(upload.name or '').lower()
        if filename.endswith('.xls') and not filename.endswith('.xlsx'):
            raise ImportValidationError('Legacy .xls files are not enabled. Please upload .xlsx.')
        if not filename.endswith(self.allowed_extensions):
            raise ImportValidationError('Invalid file format. Expected .xlsx')
        if upload.size > MAX_FILE_SIZE_BYTES:
            raise ImportValidationError('File size exceeds 10 MB limit', status_code=413)

        content = upload.read()
        try:
            upload.seek(0)
        except Exception:
            pass

        if self._contains_vba(content):
            raise ImportValidationError('Files with macros are not permitted')
        return content

    def parse_workbook(self, upload) -> ParsedWorkbook:
        content = self.validate_file(upload)
        file_hash = hashlib.sha256(content).hexdigest()
        try:
            workbook = load_workbook(
                filename=BytesIO(content),
                read_only=True,
                data_only=False,
            )
        except Exception:
            raise ImportValidationError(
                'File is corrupted or not a valid Excel workbook. Check the file format and Excel version compatibility.'
            )

        try:
            worksheet = workbook.worksheets[0]
            header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1), []))
            headers = [normalize_cell_value(cell.value) for cell in header_cells]
            missing = [header for header in REQUIRED_HEADERS if header not in headers]
            if missing:
                raise ImportValidationError(
                    f"Missing required headers: {', '.join(missing)}",
                    details=[{'missing_headers': missing}],
                )

            header_positions = {
                HEADER_TO_FIELD[header]: index
                for index, header in enumerate(headers)
                if header in HEADER_TO_FIELD
            }

            rows = []
            for excel_row in worksheet.iter_rows(min_row=2):
                if all(normalize_cell_value(cell.value) == '' for cell in excel_row):
                    continue

                if len(rows) >= MAX_ROWS:
                    raise ImportValidationError('File exceeds maximum of 1000 rows')

                row_data = {'row_number': excel_row[0].row if excel_row else len(rows) + 2}
                for field_name, index in header_positions.items():
                    cell = excel_row[index] if index < len(excel_row) else None
                    if cell is not None and (cell.data_type == 'f' or str(cell.value or '').startswith('=')):
                        raise ImportValidationError(
                            f"Row {row_data['row_number']}: Formula cells are not allowed in imported fields"
                        )
                    row_data[field_name] = sanitize_text(cell.value if cell is not None else '')
                rows.append(row_data)

            if not rows:
                raise ImportValidationError('File contains no data rows')

            return ParsedWorkbook(
                filename=str(upload.name or 'import.xlsx'),
                file_size_bytes=upload.size,
                file_hash=file_hash,
                rows=rows,
            )
        finally:
            workbook.close()

    def _contains_vba(self, content: bytes) -> bool:
        try:
            with zipfile.ZipFile(BytesIO(content)) as archive:
                return any(name.lower().endswith('vbaproject.bin') for name in archive.namelist())
        except zipfile.BadZipFile:
            return False


class RowValidator:
    def validate_rows(self, rows):
        issues: list[ValidationIssue] = []
        valid_rows = []

        for row in rows:
            row_issues = self.validate_row(row)
            issues.extend(row_issues)
            if not any(issue.level == 'error' for issue in row_issues):
                valid_rows.append(row)

        issues.extend(self.check_duplicates_in_file(rows))
        issues.extend(self.check_duplicates_in_db(rows))
        issues.extend(self.check_active_project_conflicts(rows))

        error_rows = {
            issue.row_number
            for issue in issues
            if issue.level == 'error' and issue.row_number is not None
        }
        valid_rows = [row for row in rows if row['row_number'] not in error_rows]
        return valid_rows, issues

    def validate_row(self, row):
        issues = []
        row_num = row['row_number']

        university_id = row.get('university_id', '').strip()
        title = row.get('title', '').strip()
        department = row.get('department', '').strip()
        project_type = row.get('project_type', '').strip()
        github_repo = row.get('github_repo', '').strip()
        supervisor_name = row.get('supervisor_name', '').strip()

        if not university_id:
            issues.append(self._error(row_num, 'university_id', 'University ID is required', row))
        if not title or len(title) > 255:
            issues.append(self._error(row_num, 'title', 'Project title is required and must not exceed 255 characters', row))
        if department not in VALID_DEPARTMENTS:
            issues.append(self._error(
                row_num,
                'department',
                f"Invalid department. Must be one of: {', '.join(VALID_DEPARTMENTS)}",
                row,
                error_type='invalid_value',
            ))
        if project_type not in VALID_PROJECT_TYPES:
            issues.append(self._error(
                row_num,
                'project_type',
                f"Invalid project type. Must be one of: {', '.join(VALID_PROJECT_TYPES)}",
                row,
                error_type='invalid_value',
            ))
        if github_repo and not self._valid_repo_url(github_repo):
            issues.append(self._error(row_num, 'github_repo', 'GitHub/GitLab repository must be a valid URL', row))
        if not supervisor_name:
            issues.append(self._error(row_num, 'supervisor_name', 'Supervisor name is required', row))

        return issues

    def check_duplicates_in_file(self, rows):
        issues = []
        by_university_id = defaultdict(list)
        by_title = defaultdict(list)

        for row in rows:
            university_id = row.get('university_id', '').strip()
            title = row.get('title', '').strip().lower()
            if university_id:
                by_university_id[university_id].append(row)
            if title:
                by_title[title].append(row)

        for university_id, duplicate_rows in by_university_id.items():
            if len(duplicate_rows) > 1:
                row_numbers = ', '.join(str(row['row_number']) for row in duplicate_rows)
                for row in duplicate_rows:
                    issues.append(self._error(
                        row['row_number'],
                        'university_id',
                        f'Rows {row_numbers}: Duplicate university ID {university_id} found within file',
                        row,
                        error_type='duplicate',
                    ))

        for _title_key, duplicate_rows in by_title.items():
            if len(duplicate_rows) > 1:
                title = duplicate_rows[0].get('title', '')
                row_numbers = ', '.join(str(row['row_number']) for row in duplicate_rows)
                for row in duplicate_rows:
                    issues.append(ValidationIssue(
                        row_number=row['row_number'],
                        field_name='title',
                        error_message=f"Rows {row_numbers}: Duplicate project title '{title}' found within file",
                        row_data=row,
                        level='warning',
                        error_type='duplicate',
                    ))
        return issues

    def check_duplicates_in_db(self, rows):
        issues = []
        university_ids = [row.get('university_id', '').strip() for row in rows if row.get('university_id', '').strip()]
        existing_users = {
            user.username: user
            for user in User.objects.filter(username__in=university_ids)
        }

        for row in rows:
            university_id = row.get('university_id', '').strip()
            title = row.get('title', '').strip()
            if not university_id:
                continue

            user = existing_users.get(university_id)
            if user and user.role != 'student':
                issues.append(self._error(
                    row['row_number'],
                    'university_id',
                    f'University ID {university_id} exists with non-student role',
                    row,
                    error_type='duplicate',
                ))
                continue

            if user and title and StudentIdeaProposal.objects.filter(
                student=user,
                title__iexact=title,
            ).exists():
                proposal = StudentIdeaProposal.objects.filter(student=user, title__iexact=title).first()
                created = proposal.created_at.date().isoformat() if proposal and proposal.created_at else 'unknown date'
                issues.append(self._error(
                    row['row_number'],
                    'title',
                    f"Project '{title}' for student {university_id} already exists (record ID {proposal.id}, created {created})",
                    row,
                    error_type='duplicate',
                ))

        return issues

    def check_active_project_conflicts(self, rows):
        issues = []
        university_ids = [row.get('university_id', '').strip() for row in rows if row.get('university_id', '').strip()]
        students = {
            user.username: user
            for user in User.objects.filter(username__in=university_ids, role='student')
        }

        for row in rows:
            student = students.get(row.get('university_id', '').strip())
            if not student:
                continue
            conflict = self._student_conflict_message(student)
            if conflict:
                issues.append(self._error(
                    row['row_number'],
                    'university_id',
                    conflict,
                    row,
                    error_type='active_project',
                ))
        return issues

    def _student_conflict_message(self, student):
        active_proposal_statuses = ['awaiting_members', 'pending_supervisor', 'pending_hod', 'assigned']
        active_application_statuses = ['awaiting_members', 'pending_doctor', 'pending_hod', 'registered']

        if StudentIdeaProposal.objects.filter(student=student, status__in=active_proposal_statuses).exists():
            return f'Student {student.username} already has an active proposal'
        if ProjectApplication.objects.filter(student=student, status='accepted').exists():
            return f'Student {student.username} already has an accepted project application'
        if IdeaApplication.objects.filter(student=student, status__in=active_application_statuses).exists():
            return f'Student {student.username} already has an active or registered idea application'
        if ProposalInvitation.objects.filter(
            invitee=student,
            status='accepted',
            proposal__status__in=active_proposal_statuses,
        ).exists():
            return f'Student {student.username} is already an accepted member of an active proposal'
        if TeamInvitation.objects.filter(
            invitee=student,
            status='accepted',
            application__status__in=active_application_statuses,
        ).exists():
            return f'Student {student.username} is already an accepted member of an active application'
        return ''

    def _valid_repo_url(self, value):
        parsed = urlparse(value)
        if parsed.scheme not in ('http', 'https') or not parsed.netloc:
            return False
        host = parsed.netloc.lower()
        return host == 'github.com' or host.endswith('.github.com') or host == 'gitlab.com' or host.endswith('.gitlab.com')

    def _error(self, row_number, field_name, message, row_data, *, error_type='validation'):
        return ValidationIssue(
            row_number=row_number,
            field_name=field_name,
            error_message=f'Row {row_number}: {message}',
            row_data=row_data,
            error_type=error_type,
        )


def group_issues(issues):
    grouped = defaultdict(list)
    for issue in issues:
        grouped[issue.error_type].append(issue.to_dict())
    return dict(grouped)


USERNAME_RE = re.compile(r'[^A-Za-z0-9_]+')
