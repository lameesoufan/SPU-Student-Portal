from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook

from accounts.models import User
from project_imports.constants import FIELD_HEADERS
from project_imports.services import UserMapper
from project_imports.templates import TemplateGenerator
from project_imports.validators import FileValidator, RowValidator


class ProjectImportEmailTests(TestCase):
    def _valid_row(self, **overrides):
        row = {
            'row_number': 2,
            'project_row_number': 2,
            'is_project_leader': True,
            'student_name': 'طالب تجريبي',
            'university_id': '20250001',
            'email': 'student@example.com',
            'title': 'مشروع تجريبي',
            'department': 'software_engineering',
            'supervisor_name': 'د. مشرف تجريبي',
            'project_type': 'graduation_1',
            'github_repo': 'https://github.com/example/project',
        }
        row.update(overrides)
        return row

    def test_generated_template_contains_required_email_column(self):
        content = TemplateGenerator().generate_template()
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            headers = [cell.value for cell in workbook['Projects'][1]]
        finally:
            workbook.close()
        self.assertIn(FIELD_HEADERS['email'], headers)

    def test_parser_reads_email_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([
            FIELD_HEADERS['student_name'],
            FIELD_HEADERS['university_id'],
            FIELD_HEADERS['email'],
            FIELD_HEADERS['title'],
            FIELD_HEADERS['department'],
            FIELD_HEADERS['supervisor_name'],
            FIELD_HEADERS['project_type'],
            FIELD_HEADERS['github_repo'],
        ])
        worksheet.append([
            'طالب تجريبي', '20250001', 'Student@Example.COM', 'مشروع تجريبي',
            'software_engineering', 'د. مشرف تجريبي', 'graduation_1',
            'https://github.com/example/project',
        ])
        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            'projects.xlsx', output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        parsed = FileValidator().parse_workbook(upload)
        self.assertEqual(parsed.rows[0]['email'], 'student@example.com')

    def test_email_is_required_and_must_be_valid(self):
        missing = RowValidator().validate_row(self._valid_row(email=''))
        invalid = RowValidator().validate_row(self._valid_row(email='not-an-email'))
        self.assertTrue(any(issue.field_name == 'email' for issue in missing))
        self.assertTrue(any(issue.field_name == 'email' for issue in invalid))

    def test_new_student_is_created_with_email_and_export_includes_it(self):
        User.objects.create_user(
            username='supervisor',
            password='x',
            first_name='مشرف',
            last_name='تجريبي',
            role='doctor',
            department='software_engineering',
        )
        row = self._valid_row(supervisor_name='supervisor')
        mapper = UserMapper()
        user_map = mapper.resolve_users([row])
        student = User.objects.get(username='20250001')
        self.assertEqual(student.email, 'student@example.com')

        export = mapper._build_student_credentials_export([row], user_map)
        self.assertIn('email', export['columns'])
        self.assertEqual(export['rows'][0]['email'], 'student@example.com')

    def test_existing_student_without_email_is_updated(self):
        student = User.objects.create_user(
            username='20250001',
            password='x',
            role='student',
            email='',
            department='software_engineering',
        )
        User.objects.create_user(
            username='supervisor',
            password='x',
            role='doctor',
            department='software_engineering',
        )
        mapper = UserMapper()
        mapper.resolve_users([self._valid_row(supervisor_name='supervisor')])
        student.refresh_from_db()
        self.assertEqual(student.email, 'student@example.com')

    def test_email_used_by_another_account_is_rejected(self):
        User.objects.create_user(
            username='other',
            password='x',
            role='student',
            email='student@example.com',
        )
        issues = RowValidator().check_duplicates_in_db([self._valid_row()])
        self.assertTrue(any(issue.field_name == 'email' and issue.error_type == 'duplicate' for issue in issues))
