from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import FIELD_HEADERS, REQUIRED_HEADERS, VALID_DEPARTMENTS, VALID_PROJECT_TYPES


class TemplateGenerator:
    def generate_template(self):
        workbook = Workbook()
        projects = workbook.active
        projects.title = 'Projects'
        self._create_projects_sheet(projects)
        self._create_instructions_sheet(workbook)

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)
        return output.getvalue()

    def _create_projects_sheet(self, worksheet):
        worksheet.append(REQUIRED_HEADERS)
        worksheet.append([
            'محمد أحمد',
            '20250001',
            'student1@example.com',
            'نظام إدارة مشاريع التخرج',
            'software_engineering',
            'dr_ali',
            'graduation_1',
            'https://github.com/example/spu-project',
        ])
        worksheet.append([
            'سارة خالد',
            '20250002',
            'student2@example.com',
            'تحليل ذكي للبيانات الجامعية',
            'artificial_intelligence',
            'dr_sara',
            'graduation_2',
            '',
        ])

        header_fill = PatternFill('solid', fgColor='E0E7FF')
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for column_idx in range(1, len(REQUIRED_HEADERS) + 1):
            worksheet.column_dimensions[get_column_letter(column_idx)].width = 24

        worksheet.freeze_panes = 'A2'
        worksheet.sheet_view.rightToLeft = True

    def _create_instructions_sheet(self, workbook):
        worksheet = workbook.create_sheet('Instructions')
        worksheet.sheet_view.rightToLeft = True
        rows = [
            ['Arabic Header', 'English Header', 'Description'],
            [FIELD_HEADERS['student_name'], 'student_name', 'Full student name. It will be split into first and last name.'],
            [FIELD_HEADERS['university_id'], 'university_id', 'Student university ID. This becomes the student username.'],
            [FIELD_HEADERS['email'], 'email', 'Required student email. The first-login verification code is sent to this address.'],
            [FIELD_HEADERS['title'], 'title', 'Project title. Maximum 255 characters.'],
            [FIELD_HEADERS['department'], 'department', f"One of: {', '.join(VALID_DEPARTMENTS)}"],
            [FIELD_HEADERS['supervisor_name'], 'supervisor_name', 'Doctor username or unique full/partial doctor name.'],
            [FIELD_HEADERS['project_type'], 'project_type', f"One of: {', '.join(VALID_PROJECT_TYPES)}"],
            [FIELD_HEADERS['github_repo'], 'github_repo', 'Optional GitHub or GitLab URL.'],
        ]
        for row in rows:
            worksheet.append(row)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='DCFCE7')
        worksheet.column_dimensions['A'].width = 24
        worksheet.column_dimensions['B'].width = 24
        worksheet.column_dimensions['C'].width = 90
