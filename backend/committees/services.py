"""
Business logic for the committees app — REVISED DESIGN.

Functions:
  - spawn_committee_for_template() : create exactly ONE Committee per template
  - collect_projects()              : gather projects matching (dept, project_type)
  - distribute_projects_to_committees() :
        For each (department, project_type):
          For each of the 4 committee types:
            Round-robin distribute projects across committees of that type.
            Each project appears in 4 committees (one per type).
  - copy_template()                : clone a template (1:1)
  - get_dashboard_warnings()       : check for missing chairs, undistributed, etc.
  - get_doctor_workload()          : workload per doctor
  - export_*                       : PDF & Excel
"""
from __future__ import annotations
import io
import random
from dataclasses import dataclass, field
from typing import Iterable

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

from accounts.models import DEPARTMENTS
from .models import (
    CommitteeTemplate, Committee,
    COMMITTEE_TYPE_CHOICES, PROJECT_TYPE_CHOICES,
    ALL_COMMITTEE_TYPES,
)
from projects.participation_services import get_project_participations, user_display_name


User = get_user_model()

# ── Arabic font + text shaping helpers ────────────────────────────────────────
# ReportLab does NOT do Arabic shaping or RTL flipping out of the box.
# Without these helpers, Arabic text in PDFs would either:
#   - render as boxes (if the font has no Arabic glyphs — e.g. DejaVuSans)
#   - render as disconnected letters in LTR order (with the right font but
#     without arabic_reshaper + python-bidi)
# Both helpers degrade gracefully so the PDF never crashes — it just looks
# bad if `arabic_reshaper` / `python-bidi` aren't installed.

_ARABIC_FONT_REGISTERED = None  # cached font name after first registration


def _register_arabic_font() -> str:
    """
    Register an Arabic-capable TTF font with reportlab.
    Tries multiple candidate paths across Linux / macOS / Windows; falls
    back to 'Helvetica' (which has NO Arabic glyphs) only as last resort.

    Install on Linux:   sudo apt install fonts-noto-core
    Install on Windows: download NotoSansArabic-Regular.ttf into C:\\Windows\\Fonts\\
    Or bundle in app:   committees/static/fonts/NotoSansArabic-Regular.ttf
    """
    global _ARABIC_FONT_REGISTERED
    if _ARABIC_FONT_REGISTERED is not None:
        return _ARABIC_FONT_REGISTERED

    from reportlab.pdfbase       import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    project_font = os.path.join(
        os.path.dirname(__file__), 'static', 'fonts',
        'NotoSansArabic-Regular.ttf',
    )

    candidate_paths = [
        # Linux — apt install fonts-noto-core / fonts-noto
        '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
        '/usr/share/fonts/opentype/noto/NotoSansArabic-Regular.ttf',
        '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
        # Amiri (popular Arabic serif font on Linux)
        '/usr/share/fonts/truetype/amiri/amiri-regular.ttf',
        # Project-bundled (recommended — works everywhere)
        project_font,
        # Windows — Arial has Arabic glyphs on Windows
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\NotoSansArabic-Regular.ttf',
        # macOS
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        # LAST RESORT — DejaVuSans has very weak Arabic support and will
        # likely produce boxes for many characters. Kept only so the function
        # never returns 'Helvetica' if anything else is available.
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]

    arabic_font = 'Helvetica'
    for path in candidate_paths:
        if not os.path.exists(path):
            continue
        try:
            pdfmetrics.registerFont(TTFont('ArabicFont', path))
            arabic_font = 'ArabicFont'
            break
        except Exception:
            continue

    _ARABIC_FONT_REGISTERED = arabic_font
    return arabic_font


def _ar(text) -> str:
    """
    Reshape an Arabic string for correct RTL rendering in ReportLab:
      1. arabic_reshaper connects cursive letters (mandatory for Arabic)
      2. python-bidi flips the visual order so the string displays RTL

    Returns the original text (str) unchanged if either library is missing,
    so the code never crashes — the PDF just won't look right.

    Non-Arabic text (English, numbers) passes through unaffected.
    """
    if text is None:
        return ''
    if not isinstance(text, str):
        text = str(text)
    if not text:
        return text
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except ImportError:
        # Graceful degradation — text will be wrong but won't crash.
        return text
# ── 1) Spawn ONE committee per template ───────────────────────────────────────

@transaction.atomic
def spawn_committee_for_template(template: CommitteeTemplate) -> Committee | None:
    """
    Create exactly ONE Committee instance from this template.
    Idempotent — if a committee already exists for this template, return it.

    For SINGLE mode templates: NO committee is spawned here. Committees are
    created later by `distribute_single_mode_projects()` when projects are
    distributed (4 committees per project, one per committee_type, all sharing
    the same scheduling_group UUID).
    """
    # Single mode: skip — committees created at distribution time
    if getattr(template, 'scheduling_mode', 'multi') == 'single':
        return None

    existing = template.committees.first()
    if existing:
        return existing

    members_qs = list(template.members.all())
    c = Committee.objects.create(
        template       = template,
        sequence_number= 1,
        committee_type = template.committee_type,
        department     = template.department,
        project_type   = template.project_type,
        semester       = template.semester,
        chair          = template.chair,
        status         = 'draft',
    )
    if members_qs:
        c.members.set(members_qs)
    return c


# Keep backward-compat alias for any caller still using the old name
def spawn_committees_for_template(template: CommitteeTemplate, count: int | None = None):
    """Backward-compatible wrapper — returns a list (may be empty for single mode)."""
    c = spawn_committee_for_template(template)
    return [c] if c else []


# ── 2) Collect projects for a (department, project_type) ──────────────────────

@dataclass
class CollectedProject:
    source: str          # 'IdeaApplication' | 'StudentIdeaProposal'
    id: int
    title: str
    department: str
    project_type: str | None
    supervisor_id: int | None
    supervisor_name: str
    student_id: int
    student_name: str
    team_size: int
    active_students: list[dict] = field(default_factory=list)
    inactive_students: list[dict] = field(default_factory=list)
    active_team_size: int = 0
    original_team_size: int = 0
    operational_status: str = 'active'


def collect_projects_for_template(template: CommitteeTemplate) -> list[CollectedProject]:
    """
    Collect all projects matching this template's
        (department, project_type)
    from BOTH project sources:
      - IdeaApplication.idea.department AND .project_type, status='registered'
      - StudentIdeaProposal.department  AND .project_type, status='assigned'
    """
    return _collect_projects(
        department   = template.department,
        project_type = template.project_type,
    )


def _collect_projects(department: str, project_type: str | None) -> list[CollectedProject]:
    """Lower-level collector — used by both template-based and (dept, ptype)-based flows."""
    result: list[CollectedProject] = []

    def participation_payloads(project, legacy_students):
        participations = list(get_project_participations(project))
        if not participations:
            return legacy_students, []

        active_students = []
        inactive_students = []
        for participation in participations:
            payload = {
                'id': participation.student_id,
                'name': user_display_name(participation.student),
                'university_id': participation.student.username,
                'role': participation.role,
                'is_leader': participation.role == 'leader',
                'status': participation.status,
                'designation_date': participation.status_changed_at.isoformat() if participation.status_changed_at else None,
                'reason': participation.status_reason,
            }
            if participation.status == 'active':
                active_students.append(payload)
            else:
                inactive_students.append(payload)
        return active_students, inactive_students

    # ── IdeaApplication source ────────────────────────────────────────────────
    from projects.models import IdeaApplication
    idea_apps = (
        IdeaApplication.objects
        .filter(status='registered',
                idea__department=department)
        .exclude(operational_status__in=['fully_withdrawn', 'fully_failed', 'inactive'])
        .select_related('idea', 'idea__doctor', 'student')
        .prefetch_related('invitations__invitee', 'participations__student')
    )
    for app in idea_apps:
        ptype = getattr(app, 'project_type', None) or getattr(app.idea, 'project_type', None)
        if project_type and ptype != project_type:
            continue
        legacy_students = []
        if app.student_id:
            legacy_students.append({
                'id': app.student_id,
                'name': app.student.get_full_name() or app.student.username,
                'university_id': app.student.username,
                'role': 'leader',
                'is_leader': True,
                'status': 'active',
            })
        for invitation in app.invitations.filter(status='accepted'):
            if invitation.invitee_id:
                legacy_students.append({
                    'id': invitation.invitee_id,
                    'name': invitation.invitee.get_full_name() or invitation.invitee.username,
                    'university_id': invitation.invitee.username,
                    'role': 'member',
                    'is_leader': False,
                    'status': 'active',
                })

        active_students, inactive_students = participation_payloads(app, legacy_students)
        if not active_students:
            continue
        result.append(CollectedProject(
            source          = 'IdeaApplication',
            id              = app.id,
            title           = app.idea.title,
            department      = app.idea.department,
            project_type    = ptype,
            supervisor_id   = app.idea.doctor_id,
            supervisor_name = (app.idea.doctor.get_full_name() or app.idea.doctor.username)
                              if app.idea.doctor_id else '',
            student_id      = active_students[0]['id'],
            student_name    = active_students[0]['name'],
            team_size       = len(active_students) + len(inactive_students),
            active_students = active_students,
            inactive_students = inactive_students,
            active_team_size = len(active_students),
            original_team_size = len(active_students) + len(inactive_students),
            operational_status = app.operational_status,
        ))

    # ── StudentIdeaProposal source ────────────────────────────────────────────
    from projects.models import StudentIdeaProposal
    proposals = (
        StudentIdeaProposal.objects
        .filter(status='assigned',
                department=department)
        .exclude(operational_status__in=['fully_withdrawn', 'fully_failed', 'inactive'])
        .select_related('student', 'supervisor')
        .prefetch_related('invitations__invitee', 'participations__student')
    )
    for prop in proposals:
        ptype = getattr(prop, 'project_type', None)
        if project_type and ptype != project_type:
            continue
        sup_name = ''
        if prop.supervisor_id:
            sup_name = prop.supervisor.get_full_name() or prop.supervisor.username
        legacy_students = []
        if prop.student_id:
            legacy_students.append({
                'id': prop.student_id,
                'name': prop.student.get_full_name() or prop.student.username,
                'university_id': prop.student.username,
                'role': 'leader',
                'is_leader': True,
                'status': 'active',
            })
        for invitation in prop.invitations.filter(status='accepted'):
            if invitation.invitee_id:
                legacy_students.append({
                    'id': invitation.invitee_id,
                    'name': invitation.invitee.get_full_name() or invitation.invitee.username,
                    'university_id': invitation.invitee.username,
                    'role': 'member',
                    'is_leader': False,
                    'status': 'active',
                })

        active_students, inactive_students = participation_payloads(prop, legacy_students)
        if not active_students:
            continue
        result.append(CollectedProject(
            source          = 'StudentIdeaProposal',
            id              = prop.id,
            title           = prop.title,
            department      = prop.department,
            project_type    = ptype,
            supervisor_id   = prop.supervisor_id,
            supervisor_name = sup_name,
            student_id      = active_students[0]['id'],
            student_name    = active_students[0]['name'],
            team_size       = len(active_students) + len(inactive_students),
            active_students = active_students,
            inactive_students = inactive_students,
            active_team_size = len(active_students),
            original_team_size = len(active_students) + len(inactive_students),
            operational_status = prop.operational_status,
        ))

    return result


# ── 3) Distribution: Round-Robin across committees of each type ────────────────
#
# REVISED ALGORITHM
# ─────────────────
# For each (department, project_type) combination that has projects:
#   For each of the 4 committee types (seminar_1, seminar_2, technical, final_discussion):
#     - Get all committees of that type matching the (department, project_type)
#     - If no committees of that type exist → mark all projects as undistributed for that type
#     - Otherwise: round-robin distribute the projects across these committees
#       (with wrap-around if projects > committees; extras load the first committees)
#
# Each project is therefore assigned to up to 4 committees (one per type).

def _distribution_exclusion_summary(department: str, project_type: str | None) -> dict:
    from projects.models import IdeaApplication, StudentIdeaProposal

    summary = {
        'excluded_students_total': 0,
        'excluded_failed_students': 0,
        'excluded_withdrawn_students': 0,
        'excluded_projects_zero_active': 0,
    }

    def add_project(project, ptype):
        if project_type and ptype != project_type:
            return
        participations = list(get_project_participations(project))
        if not participations:
            return
        active_count = sum(1 for p in participations if p.status == 'active')
        failed_count = sum(1 for p in participations if p.status == 'failed')
        withdrawn_count = sum(1 for p in participations if p.status == 'withdrawn')

        summary['excluded_failed_students'] += failed_count
        summary['excluded_withdrawn_students'] += withdrawn_count
        summary['excluded_students_total'] += failed_count + withdrawn_count
        if active_count == 0:
            summary['excluded_projects_zero_active'] += 1

    idea_apps = (
        IdeaApplication.objects
        .filter(status='registered', idea__department=department)
        .select_related('idea')
        .prefetch_related('participations')
    )
    for app in idea_apps:
        add_project(app, getattr(app, 'project_type', None) or getattr(app.idea, 'project_type', None))

    proposals = (
        StudentIdeaProposal.objects
        .filter(status='assigned', department=department)
        .prefetch_related('participations')
    )
    for prop in proposals:
        add_project(prop, getattr(prop, 'project_type', None))

    return summary


@dataclass
class TypeDistribution:
    """Result of distributing projects to committees of a single committee type."""
    committee_type: str
    committees_count: int
    assignments: list[dict] = field(default_factory=list)  # [{committee_id, seq, project}]
    undistributed: list[dict] = field(default_factory=list)  # projects that couldn't be placed


@dataclass
class DistributionPlan:
    template_id: int | None       # None when triggered without a specific template
    department: str
    project_type: str
    semester: str | None
    projects_count: int
    by_type: list[TypeDistribution] = field(default_factory=list)


def build_distribution_plan(template: CommitteeTemplate,
                             projects: list[CollectedProject] | None = None,
                             ) -> DistributionPlan:
    """
    Build a distribution plan for a single template's (department, project_type).
    Iterates over ALL 4 committee types and assigns projects to each type's committees.
    """
    if projects is None:
        projects = collect_projects_for_template(template)

    plan = DistributionPlan(
        template_id   = template.id,
        department    = template.department,
        project_type  = template.project_type,
        semester      = template.semester,
        projects_count= len(projects),
    )

    for ctype in ALL_COMMITTEE_TYPES:
        committees = list(
            Committee.objects
            .filter(committee_type=ctype,
                    department   =template.department,
                    project_type =template.project_type)
            .order_by('sequence_number', 'id')
        )

        td = TypeDistribution(committee_type=ctype, committees_count=len(committees))

        if not committees:
            # No committees of this type — all projects are undistributed for this type
            td.undistributed = [_project_to_dict(p) for p in projects]
        else:
            # Shuffle for fair distribution
            shuffled = list(projects)
            random.shuffle(shuffled)
            for idx, proj in enumerate(shuffled):
                committee = committees[idx % len(committees)]
                td.assignments.append({
                    'committee_id'   : committee.id,
                    'sequence_number': committee.sequence_number,
                    'project'        : _project_to_dict(proj),
                })

        plan.by_type.append(td)

    return plan


def build_distribution_plan_for_combo(department: str,
                                       project_type: str,
                                       semester: str | None = None,
                                       projects: list[CollectedProject] | None = None,
                                       ) -> DistributionPlan:
    """Same as build_distribution_plan but without a template (driven by dept+ptype)."""
    if projects is None:
        projects = _collect_projects(department, project_type)

    plan = DistributionPlan(
        template_id   = None,
        department    = department,
        project_type  = project_type,
        semester      = semester,
        projects_count= len(projects),
    )

    for ctype in ALL_COMMITTEE_TYPES:
        committees = list(
            Committee.objects
            .filter(committee_type=ctype,
                    department   =department,
                    project_type =project_type)
            .order_by('sequence_number', 'id')
        )

        td = TypeDistribution(committee_type=ctype, committees_count=len(committees))

        if not committees:
            td.undistributed = [_project_to_dict(p) for p in projects]
        else:
            shuffled = list(projects)
            random.shuffle(shuffled)
            for idx, proj in enumerate(shuffled):
                committee = committees[idx % len(committees)]
                td.assignments.append({
                    'committee_id'   : committee.id,
                    'sequence_number': committee.sequence_number,
                    'project'        : _project_to_dict(proj),
                })

        plan.by_type.append(td)

    return plan


@transaction.atomic
def apply_distribution_plan(plan: DistributionPlan) -> int:
    """
    Persist a distribution plan to the DB.
    Clears existing project assignments on the affected committees first.
    Returns the number of assignments written.
    """
    committee_ids = set()
    for td in plan.by_type:
        for a in td.assignments:
            committee_ids.add(a['committee_id'])

    committees = Committee.objects.filter(
        committee_type__in=[td.committee_type for td in plan.by_type],
        department=plan.department,
        project_type=plan.project_type,
    )
    by_id = {c.id: c for c in committees}

    # Clear existing project assignments on these committees
    for c in committees:
        c.applications.clear()
        c.proposals.clear()

    written = 0
    from projects.models import IdeaApplication, StudentIdeaProposal
    for td in plan.by_type:
        for a in td.assignments:
            c = by_id.get(a['committee_id'])
            if not c:
                continue
            p = a['project']
            if p['source'] == 'IdeaApplication':
                c.applications.add(p['id'])
            else:
                c.proposals.add(p['id'])
            written += 1

    return written


@transaction.atomic
def distribute_projects_to_committees(template_ids: list[int] | None = None,
                                       semester: str | None = None,
                                       dry_run: bool = False,
                                       scheduling_mode: str = 'multi',
                                       ) -> dict:
    """Distribute projects without deleting unrelated committees.

    The redistribution scope is derived from the selected templates as
    (department, project_type). Project and committee semester values are not
    used to restrict redistribution; all current projects in the combo are
    included. Committees in other departments or project types remain untouched.
    """
    selected_templates = CommitteeTemplate.objects.all()
    if template_ids:
        selected_templates = selected_templates.filter(id__in=template_ids)
    if semester:
        selected_templates = selected_templates.filter(semester=semester)

    # Materialize once: the same target set is used to derive safe scopes and
    # to avoid queryset changes while committees are deleted/recreated.
    selected_templates = list(
        selected_templates.select_related('chair').prefetch_related('members')
    )
    target_scopes = {
        (tmpl.department, tmpl.project_type)
        for tmpl in selected_templates
    }

    deleted_count = 0
    single_results = []
    plans: list[DistributionPlan] = []
    total_distributed = 0
    total_undistributed = 0
    exclusion_totals = {
        'excluded_students_total': 0,
        'excluded_failed_students': 0,
        'excluded_withdrawn_students': 0,
        'excluded_projects_zero_active': 0,
    }

    if scheduling_mode == 'single':
        # The single-mode worker already deletes only its exact
        # (department, project_type, semester) scope.
        selected_template_ids_by_scope = {}
        for tmpl in selected_templates:
            scope = (tmpl.department, tmpl.project_type)
            selected_template_ids_by_scope.setdefault(scope, []).append(tmpl.id)

        for dept, ptype in sorted(target_scopes):
            result = distribute_single_mode_projects(
                department=dept,
                project_type=ptype,
                semester=None,
                template_ids=selected_template_ids_by_scope.get((dept, ptype)),
                dry_run=dry_run,
            )
            single_results.append(result)
            deleted_count += result.get('previous_committees_deleted', 0)

    else:
        # MULTI MODE: rebuild only the selected scopes. When template_ids select
        # one template from a scope, all templates in that same scope are
        # respawned so valid sibling committees are not accidentally lost.
        for dept, ptype in sorted(target_scopes):
            scope_templates = [
                tmpl for tmpl in selected_templates
                if tmpl.department == dept and tmpl.project_type == ptype
            ]

            if not dry_run:
                scoped_committees = Committee.objects.filter(
                    department=dept,
                    project_type=ptype,
                )
                deleted_count += scoped_committees.count()
                scoped_committees.delete()

                for tmpl in scope_templates:
                    members_qs = list(tmpl.members.all())
                    c = Committee.objects.create(
                        template=tmpl,
                        sequence_number=1,
                        committee_type=tmpl.committee_type,
                        department=tmpl.department,
                        project_type=tmpl.project_type,
                        semester=tmpl.semester,
                        chair=tmpl.chair,
                        status='draft',
                        discussion_duration=getattr(tmpl, 'discussion_duration', None) or 15,
                    )
                    if members_qs:
                        c.members.set(members_qs)

            summary = _distribution_exclusion_summary(dept, ptype)
            for key, value in summary.items():
                exclusion_totals[key] += value

            projects = _collect_projects(dept, ptype)
            plan = build_distribution_plan_for_combo(
                dept,
                ptype,
                semester=None,
                projects=projects,
            )

            if not dry_run:
                apply_distribution_plan(plan)

            for td in plan.by_type:
                total_distributed += len(td.assignments)
                total_undistributed += len(td.undistributed)
            plans.append(plan)

    single_committees_created = sum(
        r.get('committees_created', 0) for r in single_results
    )
    single_projects_distributed = sum(
        r.get('projects_count', 0) for r in single_results
    )

    return {
        'processed_templates': len(selected_templates),
        'processed_scopes': len(target_scopes),
        'previous_committees_deleted': deleted_count,
        'distributed_projects': total_distributed + single_projects_distributed,
        'undistributed_projects': total_undistributed,
        'exclusions': exclusion_totals,
        'plans': [_plan_to_dict(p) for p in plans],
        'single_mode_results': single_results,
        'single_mode_committees_created': single_committees_created,
        'dry_run': dry_run,
        'executed_at': timezone.now().isoformat(),
    }


def _project_to_dict(p: CollectedProject) -> dict:
    return {
        'source':          p.source,
        'id':              p.id,
        'title':           p.title,
        'department':      p.department,
        'project_type':    p.project_type,
        'supervisor_id':   p.supervisor_id,
        'supervisor_name': p.supervisor_name,
        'student_id':      p.student_id,
        'student_name':    p.student_name,
        'team_size':       p.team_size,
        'active_students': p.active_students,
        'inactive_students': p.inactive_students,
        'active_team_size': p.active_team_size,
        'original_team_size': p.original_team_size,
        'operational_status': p.operational_status,
    }


def _type_dist_to_dict(td: TypeDistribution) -> dict:
    return {
        'committee_type'  : td.committee_type,
        'committees_count': td.committees_count,
        'assignments'     : td.assignments,
        'undistributed'   : td.undistributed,
    }


def _plan_to_dict(plan: DistributionPlan) -> dict:
    return {
        'template_id'     : plan.template_id,
        'department'      : plan.department,
        'project_type'    : plan.project_type,
        'semester'        : plan.semester,
        'projects_count'  : plan.projects_count,
        'by_type'         : [_type_dist_to_dict(td) for td in plan.by_type],
    }


# ── 4) Copy a template ────────────────────────────────────────────────────────

@transaction.atomic
def copy_template(source: CommitteeTemplate,
                  copy_doctors: bool = True,
                  new_committee_type: str | None = None,
                  new_department: str | None = None,
                  new_project_type: str | None = None,
                  new_semester: str | None = None,
                  committees_count: int | None = None,  # accepted for backward-compat, ignored
                  created_by: User | None = None,
                  ) -> CommitteeTemplate:
    """
    Clone a template. The new template creates exactly ONE committee (handled by the
    viewset's perform_create). Committees_count parameter is ignored in the new design.
    """
    new = CommitteeTemplate.objects.create(
        name           = (source.name or '') + ' (نسخة)',
        committee_type = new_committee_type or source.committee_type,
        department     = new_department    or source.department,
        project_type   = new_project_type  or source.project_type,
        semester       = new_semester      if new_semester is not None else source.semester,
        is_approved    = False,
        created_by     = created_by or source.created_by,
    )
    if copy_doctors:
        if source.chair_id:
            new.chair = source.chair
            new.save(update_fields=['chair'])
        new.members.set(list(source.members.all()))
    return new


# ── 5) Warnings ───────────────────────────────────────────────────────────────

def get_dashboard_warnings(semester: str | None = None) -> list[dict]:
    """
    Return a list of warnings (NON-BLOCKING — Dean is informed, never forced).
    Each warning: {level: 'error'|'warn'|'info', code: str, message: str, related_id: int|None}
    """
    warnings: list[dict] = []

    qs = Committee.objects.all()
    if semester:
        qs = qs.filter(semester=semester)

    # ── Committees without a chair ────────────────────────────────────────────
    for c in qs.filter(chair__isnull=True):
        warnings.append({
            'level':      'error',
            'code':       'no_chair',
            'message':    f'اللجنة "{c}" ليس لديها رئيس. يرجى تعيين رئيس قبل التوزيع.',
            'related_id': c.id,
        })

    # ── Projects without matching committees (per committee type) ──────────────
    # For each (department, project_type) that has projects, check whether committees
    # exist for EACH of the 4 committee types. If a type is missing, warn.
    from projects.models import IdeaApplication, StudentIdeaProposal

    dept_pt_with_projects = set()
    active_project_statuses = ['active', 'partial_team', 'solo']
    for app in IdeaApplication.objects.filter(
        status='registered',
        operational_status__in=active_project_statuses,
    ).select_related('idea'):
        ptype = getattr(app.idea, 'project_type', None)
        dept_pt_with_projects.add((app.idea.department, ptype))
    for prop in StudentIdeaProposal.objects.filter(
        status='assigned',
        operational_status__in=active_project_statuses,
    ):
        ptype = getattr(prop, 'project_type', None)
        dept_pt_with_projects.add((prop.department, ptype))

    existing_combos = set(
        Committee.objects.values_list('committee_type', 'department', 'project_type')
    )

    for dept, ptype in dept_pt_with_projects:
        for ctype in ALL_COMMITTEE_TYPES:
            if (ctype, dept, ptype) not in existing_combos:
                from .models import COMMITTEE_TYPE_AR, DEPARTMENT_AR, PROJECT_TYPE_AR
                ct_label  = COMMITTEE_TYPE_AR.get(ctype, ctype)
                dep_label = DEPARTMENT_AR.get(dept, dept)
                pt_label  = PROJECT_TYPE_AR.get(ptype, ptype) if ptype else '—'
                warnings.append({
                    'level':   'warn',
                    'code':    'missing_committee_type',
                    'message': f'لا توجد لجان من نوع "{ct_label}" للتركيبة ({dep_label} - {pt_label}). '
                               f'المشاريع المطابقة لن تُوزَّع على هذا النوع.',
                    'related_id': None,
                })

    # ── Doctor overload ───────────────────────────────────────────────────────
    workload = get_doctor_workload(semester=semester)
    for w in workload:
        if w['total_committees'] >= 6:
            warnings.append({
                'level':      'warn',
                'code':       'doctor_overload',
                'message':    f'الدكتور "{w["doctor_name"]}" في {w["total_committees"]} لجنة (عبء مرتفع).',
                'related_id': w['doctor_id'],
            })

    # ── Draft committees not yet scheduled ────────────────────────────────────
    unscheduled = qs.filter(status='draft').count()
    if unscheduled:
        warnings.append({
            'level':   'info',
            'code':    'unscheduled',
            'message': f'{unscheduled} لجنة بحالة "مسودة" — بانتظار تحديد موعد الانعقاد.',
            'related_id': None,
        })

    return warnings


# ── 6) Doctor workload ────────────────────────────────────────────────────────

def get_doctor_workload(semester: str | None = None) -> list[dict]:
    """
    Return list of doctors with their committee counts.
    workload_level: low (≤2), med (3-5), high (≥6)
    """
    qs = Committee.objects.all()
    if semester:
        qs = qs.filter(semester=semester)

    doctors: dict[int, dict] = {}
    for c in qs:
        if c.chair_id:
            d = doctors.setdefault(c.chair_id, {
                'doctor_id':        c.chair_id,
                'doctor_name':      c.chair.get_full_name() or c.chair.username,
                'department_ar':    _dept_ar(c.chair.department),
                'chaired_count':    0,
                'member_count':     0,
                'total_committees': 0,
            })
            d['chaired_count']    += 1
            d['total_committees'] += 1
        for m in c.members.all():
            d = doctors.setdefault(m.id, {
                'doctor_id':        m.id,
                'doctor_name':      m.get_full_name() or m.username,
                'department_ar':    _dept_ar(m.department),
                'chaired_count':    0,
                'member_count':     0,
                'total_committees': 0,
            })
            d['member_count']     += 1
            d['total_committees'] += 1

    for d in doctors.values():
        t = d['total_committees']
        d['workload_level'] = 'low' if t <= 2 else ('med' if t <= 5 else 'high')

    # Sort by total_committees desc
    return sorted(doctors.values(), key=lambda x: -x['total_committees'])


def _dept_ar(dept: str | None) -> str:
    if not dept:
        return ''
    from .models import DEPARTMENT_AR
    return DEPARTMENT_AR.get(dept, dept)


# ── 7) Export: PDF + Excel ────────────────────────────────────────────────────

def export_committees_pdf(semester: str | None = None) -> bytes:
    """
    Generate a PDF report of all committees (and their projects/doctors).
    Uses reportlab.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import mm
    from reportlab.lib           import colors
    from reportlab.platypus      import (SimpleDocTemplate, Paragraph, Spacer,
                                          Table, TableStyle, PageBreak)
    from reportlab.pdfbase       import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Register an Arabic-capable font (uses cached registration).
    # See _register_arabic_font() for the list of candidates.
    arabic_font = _register_arabic_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             topMargin=15*mm, bottomMargin=15*mm,
                             leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleX', parent=styles['Title'],
                                  fontName=arabic_font, fontSize=18, spaceAfter=12)
    h2_style    = ParagraphStyle('H2X', parent=styles['Heading2'],
                                  fontName=arabic_font, fontSize=13, spaceAfter=6,
                                  textColor=colors.HexColor('#3730a3'))
    body_style  = ParagraphStyle('BodyX', parent=styles['Normal'],
                                  fontName=arabic_font, fontSize=10)

    story = []
    story.append(Paragraph(_ar('تقرير اللجان - نظام إدارة مشاريع التخرج'), title_style))
    if semester:
        story.append(Paragraph(_ar(f'الفصل: {semester}'), body_style))
    story.append(Spacer(1, 8))

    qs = Committee.objects.all()
    if semester:
        qs = qs.filter(semester=semester)
    qs = qs.order_by('committee_type', 'department', 'sequence_number')

    for c in qs:
        story.append(Paragraph(_ar(f'{c} — {c.template.display_name() if c.template_id else ""}'), h2_style))

        # Doctors table
        doctors = c.get_all_doctors()
        doc_rows = [[_ar('الاسم'), _ar('الدور'), _ar('القسم')]]
        for d in doctors:
            doc_rows.append([_ar(d['name']),
                              _ar('رئيس') if d['role'] == 'chair' else _ar('عضو'),
                              _ar(_dept_ar(d['department']))])
        t = Table(doc_rows, colWidths=[60*mm, 25*mm, 50*mm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), arabic_font),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#eef2ff')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(t)
        story.append(Spacer(1, 4))

        # Projects table
        projs = c.get_all_projects()
        if projs:
            proj_rows = [[_ar('المصدر'), _ar('العنوان'), _ar('المشرف'), _ar('الطالب')]]
            for p in projs:
                supervisors_text = ', '.join(
                    supervisor.get('name', '')
                    for supervisor in p.get('supervisors', [])
                    if supervisor.get('name')
                ) or '—'
                # Show ONLY active students
                active_students_data = p.get('active_students', [])
                students_text = ', '.join(
                    student.get('name', '')
                    for student in active_students_data
                    if student.get('name')
                ) or '—'
                proj_rows.append([
                    _ar(p['source']),
                    _ar(p['title']),
                    _ar(supervisors_text),
                    _ar(students_text),
                ])
            t2 = Table(proj_rows, colWidths=[40*mm, 80*mm, 50*mm, 50*mm])
            t2.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,-1), arabic_font),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dcfce7')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(t2)
        else:
            story.append(Paragraph(_ar('(لا توجد مشاريع موزعة بعد)'), body_style))

        story.append(Spacer(1, 10))

    doc.build(story)
    return buf.getvalue()


def export_committees_excel(semester: str | None = None) -> bytes:
    """
    Generate an .xlsx workbook with one sheet per committee type.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    qs = Committee.objects.all()
    if semester:
        qs = qs.filter(semester=semester)

    # Group by committee_type
    by_type: dict[str, list[Committee]] = {}
    for c in qs:
        by_type.setdefault(c.committee_type, []).append(c)

    headers_fill = PatternFill('solid', fgColor='4F46E5')
    headers_font = Font(bold=True, color='FFFFFF', size=11)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right  = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ctype, committees in by_type.items():
        ws = wb.create_sheet(title=ctype[:30])
        headers = [
            'رقم اللجنة', 'نوع اللجنة', 'القسم', 'نوع المشروع', 'الفصل',
            'الرئيس', 'الأعضاء',
            'عدد المشاريع', 'المشاريع (العناوين)',
            'التاريخ', 'الوقت', 'الموقع', 'الحالة',
        ]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = headers_fill
            cell.font = headers_font
            cell.alignment = align_center
            cell.border = border

        from .models import COMMITTEE_TYPE_AR, DEPARTMENT_AR, PROJECT_TYPE_AR
        for c in committees:
            doctors = c.get_all_doctors()
            chair_name = next((d['name'] for d in doctors if d['role'] == 'chair'), '—')
            member_names = '، '.join(d['name'] for d in doctors if d['role'] == 'member')
            projs = c.get_all_projects()
            proj_titles = ' | '.join(p['title'] for p in projs)

            # Prefer CP-SAT scheduled fields; fall back to legacy fields.
            sched_date  = c.scheduled_start.strftime('%Y-%m-%d') if c.scheduled_start \
                else (c.date.strftime('%Y-%m-%d') if c.date else '')
            sched_start = c.scheduled_start.strftime('%H:%M') if c.scheduled_start \
                else (c.start_time.strftime('%H:%M') if c.start_time \
                else (c.time.strftime('%H:%M') if c.time else ''))
            sched_loc   = c.room.name if (c.room_id and c.room) \
                else (c.location if c.location else '')

            row = [
                f'{c.sequence_number:03d}',
                COMMITTEE_TYPE_AR.get(c.committee_type, c.committee_type),
                DEPARTMENT_AR.get(c.department, c.department),
                PROJECT_TYPE_AR.get(c.project_type, c.project_type),
                c.semester,
                chair_name,
                member_names,
                len(projs),
                proj_titles,
                sched_date,
                sched_start,
                sched_loc,
                c.status,
            ]
            ws.append(row)
            r = ws.max_row
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=r, column=col_idx).alignment = align_right
                ws.cell(row=r, column=col_idx).border = border

        # Column widths
        widths = [10, 15, 18, 14, 14, 22, 30, 12, 60, 14, 10, 18, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    if not wb.sheetnames:
        ws = wb.create_sheet(title='لا يوجد بيانات')
        ws.append(['لا توجد لجان لعرضها'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export Projects Assignment to Excel ───────────────────────────────────────

def export_projects_assignment_excel(semester: str | None = None) -> bytes:
    """
    Generate an Excel file for Projects Assignment table.
    Shows: Student, Project, Supervisor, Committee, Members, Date, Time, Location
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'توزيع المشاريع'

    # Styling
    headers_fill = PatternFill('solid', fgColor='667EEA')
    headers_font = Font(bold=True, color='FFFFFF', size=12)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right  = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    headers = [
        '#', 'الطلاب', 'المشروع', 'المشرفين', 
        'اللجنة', 'نوع اللجنة', 'القسم',
        'أعضاء اللجنة', 'التاريخ', 'وقت بداية المناقشة', 'وقت نهاية المناقشة', 'المكان'
    ]
    ws.append(headers)
    
    # Style headers
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = headers_fill
        cell.font = headers_font
        cell.alignment = align_center
        cell.border = border

    # Get committees
    committees_qs = Committee.objects.all()
    if semester:
        committees_qs = committees_qs.filter(semester=semester)

    from .models import COMMITTEE_TYPE_AR, DEPARTMENT_AR, PROJECT_TYPE_AR

    # Build data rows
    row_num = 1
    for committee in committees_qs:
        # Get all members
        all_doctors = committee.get_all_doctors()
        members_text = '\n'.join([
            f"{'👤 ' if doc['role'] == 'chair' else '• '}{doc['name']}"
            for doc in all_doctors
        ])
        
        # Get all projects
        projects = committee.get_all_projects()
        
        # Calculate project times
        project_times = committee.calculate_project_times()
        times_map = {}
        for pt in project_times:
            key = f"{pt['project_source']}-{pt['project_id']}"
            times_map[key] = {
                'scheduled_start': pt['start_time'],
                'scheduled_end': pt['end_time'],
            }
        
        for project in projects:
            row_num += 1
            
            # Format ONLY ACTIVE students (team members)
            students_data = project.get('active_students', [])
            if students_data:
                students_text = '\n'.join([
                    f"{'👤 ' if student.get('is_leader') else '• '}{student['name']}"
                    for student in students_data
                ])
            else:
                students_text = '—'
            
            # Format all supervisors
            supervisors_data = project.get('supervisors', [])
            if supervisors_data:
                supervisors_text = '\n'.join([
                    f"{'👤 ' if supervisor.get('is_main') else '• '}{supervisor['name']}"
                    for supervisor in supervisors_data
                ])
            else:
                supervisors_text = '—'
            
            # Get calculated times for this project
            key = f"{project['source']}-{project['id']}"
            scheduled_start = times_map.get(key, {}).get('scheduled_start', '—')
            scheduled_end = times_map.get(key, {}).get('scheduled_end', '—')
            
            row_data = [
                row_num - 1,  # Number
                students_text,  # All team members (leader marked with 👤)
                project.get('title', '—'),
                supervisors_text,  # All supervisors (main marked with 👤)
                f"{COMMITTEE_TYPE_AR.get(committee.committee_type, committee.committee_type)} - {DEPARTMENT_AR.get(committee.department, committee.department)}",
                COMMITTEE_TYPE_AR.get(committee.committee_type, committee.committee_type),
                DEPARTMENT_AR.get(committee.department, committee.department),
                members_text if members_text else '—',
                (committee.scheduled_start.strftime('%Y-%m-%d') if committee.scheduled_start else (committee.date.strftime('%Y-%m-%d') if committee.date else '—')),
                scheduled_start,  # Calculated start time for this project
                scheduled_end,    # Calculated end time for this project
                (committee.room.name if (committee.room_id and committee.room) else (committee.location if committee.location else '—')),
            ]
            ws.append(row_data)
            
            # Style cells
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = align_right
                cell.border = border

    # Column widths
    widths = [6, 20, 35, 20, 30, 15, 18, 35, 14, 16, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze first row
    ws.freeze_panes = 'A2'

    # If no data
    if ws.max_row == 1:
        ws.append(['لا توجد مشاريع موزعة', '', '', '', '', '', '', '', '', '', '', ''])

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Single-mode distribution ─────────────────────────────────────────────────
#
# In single mode, each CommitteeTemplate represents a unified committee that
# will evaluate a project in 4 sessions (one per committee_type). When the
# dean distributes projects:
#   1. Collect all single-mode templates for (dept × project_type)
#   2. Collect all active projects for (dept × project_type)
#   3. Round-robin assign projects to templates
#   4. For each (project, template) assignment, create 4 Committee instances:
#      - One per committee_type (seminar_1, seminar_2, technical, final_discussion)
#      - All with the same scheduling_group UUID (so they're linked)
#      - All with the same chair + members from the template
#      - Each gets the project assigned (applications.add or proposals.add)
#   5. Clear previous single-mode committees for these (dept × ptype × semester)
#      before creating new ones (idempotent re-distribution)

import uuid as _uuid
from .models import ALL_COMMITTEE_TYPES


@transaction.atomic
def distribute_single_mode_projects(
    *,
    department: str,
    project_type: str,
    semester: str | None = None,
    template_ids: list[int] | None = None,
    dry_run: bool = False,
) -> dict:
    """Distribute projects to single-mode templates for a (dept × ptype × semester).

    For each project assigned to a single-mode template, creates 4 Committee
    instances (one per committee_type) sharing the same scheduling_group UUID.

    Returns a summary dict.
    """
    # Collect ALL templates for this combo (scheduling_mode is decided at
    # distribute time, NOT stored on the template anymore)
    templates_qs = CommitteeTemplate.objects.filter(
        department=department,
        project_type=project_type,
    )
    if semester:
        templates_qs = templates_qs.filter(semester=semester)
    if template_ids:
        templates_qs = templates_qs.filter(id__in=template_ids)

    # Stable order makes repeated distributions predictable while still
    # balancing projects round-robin across every selected formation.
    templates = list(
        templates_qs
        .select_related('chair')
        .prefetch_related('members')
        .order_by('id')
    )
    if not templates:
        return {
            'department': department,
            'project_type': project_type,
            'semester': semester,
            'mode': 'single',
            'templates_count': 0,
            'projects_count': 0,
            'committees_created': 0,
            'message': 'No single-mode templates found for this combo.',
        }

    # Collect active projects for this combo
    projects = _collect_projects(department, project_type)
    if not projects:
        return {
            'department': department,
            'project_type': project_type,
            'semester': semester,
            'mode': 'single',
            'templates_count': len(templates),
            'projects_count': 0,
            'committees_created': 0,
            'message': 'No active projects to distribute.',
        }

    # If not dry_run: clear ALL previous committees for this combo
    # (dept × ptype × semester), regardless of which template they came from.
    # This ensures a clean slate — no stale committees from prior runs.
    if not dry_run:
        previous_committees = Committee.objects.filter(
            department=department,
            project_type=project_type,
        )
        deleted_count = previous_committees.count()
        previous_committees.delete()
    else:
        deleted_count = 0

    # SINGLE mode keeps the same doctors for the four sessions of one
    # project, but different projects are balanced across all selected
    # formations using round-robin.
    assignments = []
    committees_created = 0
    global_seq = 0  # running counter across all committees
    from projects.models import IdeaApplication, StudentIdeaProposal

    for idx, project in enumerate(projects):
        # Project 0 -> template 0, project 1 -> template 1, then wrap around.
        # All four committees created for this project use this same template.
        template = templates[idx % len(templates)]
        group_uuid = _uuid.uuid4()

        # Create 4 committees (one per committee_type)
        for ctype_idx, ctype in enumerate(ALL_COMMITTEE_TYPES):
            global_seq += 1
            c = Committee.objects.create(
                template=template,
                sequence_number=global_seq,
                committee_type=ctype,
                department=department,
                project_type=project_type,
                semester=template.semester,
                chair=template.chair,
                status='draft',
                scheduling_group=group_uuid,
                discussion_duration=getattr(template, 'discussion_duration', None),
            )
            # Set members
            member_ids = [m.id for m in template.members.all()]
            if member_ids:
                c.members.set(member_ids)
            # Assign the project
            if project.source == 'IdeaApplication':
                c.applications.add(project.id)
            else:
                c.proposals.add(project.id)
            committees_created += 1

        assignments.append({
            'project_source': project.source,
            'project_id': project.id,
            'project_title': project.title,
            'template_id': template.id,
            'template_name': template.display_name(),
            'scheduling_group': str(group_uuid),
        })

    template_loads = {template.id: 0 for template in templates}
    for assignment in assignments:
        template_loads[assignment['template_id']] += 1

    return {
        'department': department,
        'project_type': project_type,
        'semester': semester,
        'mode': 'single',
        'templates_count': len(templates),
        'template_loads': template_loads,
        'projects_count': len(projects),
        'committees_created': committees_created,
        'previous_committees_deleted': deleted_count,
        'assignments': assignments,
        'dry_run': dry_run,
    }
